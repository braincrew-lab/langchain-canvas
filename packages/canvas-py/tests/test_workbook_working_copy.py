"""An uploaded workbook is edited through its working copy.

The upload under ``sources/`` is read-only; a grid drawn straight from it
could be typed into but never saved, and an agent refused at the source
rebuilt the table without its formatting. The copy is the file both work on.
"""

from __future__ import annotations

import io
import json
import sys
from typing import Any

import pytest

sys.path.insert(0, __file__.rsplit("/", 1)[0])
from test_tools import _runtime  # noqa: E402

from langchain_canvas.replay import (  # noqa: E402
    encode_artifact,
    encode_table,  # noqa: E402
    hydrate_events,
    source_preview_events,
    workbook_working_copy,
    working_copy_path,
)
from langchain_canvas.store import InMemoryCanvasStore  # noqa: E402
from langchain_canvas.tools import create_canvas_tools, create_table_tools  # noqa: E402


def _workbook() -> bytes:
    openpyxl = pytest.importorskip("openpyxl")
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "원장"
    sheet.append(["상품", "단가", "수량", "금액"])
    sheet.append(["A", 100, 2, "=B2*C2"])
    sheet["A1"].font = openpyxl.styles.Font(bold=True)
    out = io.BytesIO()
    book.save(out)
    return out.getvalue()


def _uploaded(store: InMemoryCanvasStore) -> str:
    commit = store.write_bytes("t1", "sources/book.xlsx", _workbook(), "Upload", actor="human")
    return commit.revision


def test_the_copy_lands_at_the_root_with_the_sheet_state() -> None:
    store = InMemoryCanvasStore()
    _uploaded(store)
    result = workbook_working_copy(store, "t1", "sources/book.xlsx", actor="human")
    assert result is not None
    path, events = result
    assert path == "book.table.json" == working_copy_path("sources/book.xlsx")
    envelope = json.loads(store.read("t1", path).content)
    assert envelope["type"] == "table"
    assert envelope["data"]["sheet"], "the grid state (fonts, formulas) came across"
    assert envelope["data"]["rows"][0]["금액"] == "=B2*C2"  # formulas as source text
    assert [e["type"] for e in events][:1] == ["canvas.create"]


def test_the_copy_is_never_overwritten_and_an_unreadable_book_makes_none() -> None:
    store = InMemoryCanvasStore()
    _uploaded(store)
    assert workbook_working_copy(store, "t1", "sources/book.xlsx", actor="human") is not None
    mine = '{"type":"table","title":"mine","data":{}}'
    store.write("t1", "book.table.json", mine, "edit", actor="human")
    assert workbook_working_copy(store, "t1", "sources/book.xlsx", actor="human") is None
    assert json.loads(store.read("t1", "book.table.json").content)["title"] == "mine"
    store.write_bytes("t1", "sources/broken.xlsx", b"PK-nope", "Upload", actor="human")
    assert workbook_working_copy(store, "t1", "sources/broken.xlsx", actor="human") is None


def test_with_a_copy_the_upload_shows_as_a_card_not_a_second_grid() -> None:
    """Two grids meant the person edited the one that could not save."""
    store = InMemoryCanvasStore()
    revision = _uploaded(store)
    before = source_preview_events(
        store, "t1", "sources/book.xlsx", is_new=True, revision=revision, description="Upload"
    )
    assert next(e for e in before if e["type"] == "canvas.create")["artifact"]["type"] == "table"
    workbook_working_copy(store, "t1", "sources/book.xlsx", actor="human")
    after = source_preview_events(
        store, "t1", "sources/book.xlsx", is_new=True, revision=revision, description="Upload"
    )
    assert next(e for e in after if e["type"] == "canvas.create")["artifact"]["type"] == "file"
    kinds = {
        e["artifact"]["id"]: e["artifact"]["type"]
        for e in hydrate_events(store, "t1")
        if e["type"] == "canvas.create"
    }
    assert kinds == {"sources/book.xlsx": "file", "book.table.json": "table"}


def test_writing_to_the_upload_names_the_copy() -> None:
    store = InMemoryCanvasStore()
    revision = _uploaded(store)
    workbook_working_copy(store, "t1", "sources/book.xlsx", actor="human")
    rt = _runtime(thread_id="t1")
    canvas = {t.name: t for t in create_canvas_tools(store)}
    table = {t.name: t for t in create_table_tools(store)}
    refused = canvas["write_canvas"].func(
        path="sources/book.xlsx", content="x", description="d", runtime=rt
    )
    assert "book.table.json" in refused and "write_table_cells" in refused
    refused = table["write_table_cells"].func(
        path="sources/book.xlsx", sheet="s0", cells={"B2": 1}, description="d",
        revision=revision, runtime=rt,
    )
    assert "book.table.json" in refused
    read = canvas["read_canvas"].func(path="sources/book.xlsx", runtime=rt)
    text = read if isinstance(read, str) else read[0]["text"]
    assert "Editable working copy: book.table.json" in text


def _write(store: InMemoryCanvasStore, content: str, path: str = "t.table.json", **kw: Any) -> str:
    tools = {t.name: t for t in create_canvas_tools(store)}
    return tools["write_canvas"].func(
        path=path, content=content, description="d", runtime=_runtime(thread_id="t1"), **kw
    )


def test_a_table_with_keys_outside_data_is_refused_with_the_shape() -> None:
    """What an agent wrote in a run: the grid crashed on a missing `columns`."""
    store = InMemoryCanvasStore()
    reply = _write(store, json.dumps({"type": "table", "sheet": [{"name": "s"}]}))
    assert reply.startswith("Error: t.table.json was not saved")
    assert '"sheet" must sit inside "data"' in reply
    assert store.list_files("t1") == []


def test_a_table_is_normalised_so_columns_and_rows_always_exist() -> None:
    store = InMemoryCanvasStore()
    only_sheet = {"type": "table", "data": {"sheet": [{"name": "s", "celldata": []}]}}
    reply = _write(store, json.dumps(only_sheet))
    assert reply.startswith("Wrote t.table.json"), reply
    data = json.loads(store.read("t1", "t.table.json").content)["data"]
    assert data["columns"] == [] and data["rows"] == [] and data["sheet"]


def test_a_rewrite_that_would_drop_the_persons_grid_state_is_refused() -> None:
    store = InMemoryCanvasStore()
    _uploaded(store)
    workbook_working_copy(store, "t1", "sources/book.xlsx", actor="human")
    revision = store.read("t1", "book.table.json").revision
    reply = _write(
        store,
        json.dumps({"type": "table", "data": {"columns": [{"key": "a"}], "rows": [{"a": 1}]}}),
        path="book.table.json",
        revision=revision,
    )
    assert reply.startswith("Error: book.table.json was not saved")
    assert "write_table_cells" in reply
    assert json.loads(store.read("t1", "book.table.json").content)["data"]["sheet"]


def test_a_table_that_is_not_json_or_not_a_table_is_refused() -> None:
    store = InMemoryCanvasStore()
    assert "not valid JSON" in _write(store, "a,b\n1,2")
    bad = json.dumps({"type": "table", "data": {"columns": "no"}})
    assert "was not saved" in _write(store, bad)
    assert store.list_files("t1") == []


def test_encode_artifact_still_projects_rows_from_the_sheet() -> None:
    """The normalisation path must not bypass the sheet -> rows projection."""
    store = InMemoryCanvasStore()
    _uploaded(store)
    workbook_working_copy(store, "t1", "sources/book.xlsx", actor="human")
    envelope = json.loads(store.read("t1", "book.table.json").content)
    again = encode_artifact(envelope, "book.table.json")
    assert json.loads(again)["data"]["rows"] == envelope["data"]["rows"]


# --- formulas get their values at save time --------------------------------------------

# A sheets-capable stand-in evaluator: XLOOKUP fails, everything else is 42.
_GRID_EVALUATOR = (
    sys.executable,
    "-c",
    (
        "import json,sys; p=json.load(sys.stdin); r=[]\n"
        "for i,s in enumerate(p.get('sheets') or []):\n"
        "    for cell in s.get('celldata') or []:\n"
        "        v=cell.get('v')\n"
        "        f=v.get('f') if isinstance(v,dict) else None\n"
        "        if f: r.append({'sheet': i, 'r': cell['r'], 'c': cell['c'],\n"
        "                        'formula': f, 'value': '#ERR' if 'XLOOKUP' in f else 42})\n"
        "print(json.dumps({'results': r}))"
    ),
)


def _seeded_store() -> InMemoryCanvasStore:
    """A working-copy-shaped table: one imported formula with its cached value."""
    store = InMemoryCanvasStore()
    data = {
        "columns": [{"key": "a"}],
        "rows": [{"a": 1}],
        "sheet": [{"name": "Sheet1", "celldata": [
            {"r": 0, "c": 7, "v": {"v": 5}},
            {"r": 0, "c": 8, "v": {"v": 14300}},
            {"r": 5, "c": 0, "v": {"v": 999, "f": "=SUM(H1:I1)", "m": "999"}},
        ]}],
    }
    store.write("t1", "book.table.json", encode_table("Book", data), "copy", actor="agent")
    return store


def test_a_written_formula_lands_with_its_value_and_dependents_refresh() -> None:
    """Measured (thread a87118dc): the agent's `=ROUND(H2*I2*2,0)` stored only
    `f`; the grid showed a blank cell, and cells depending on a changed value
    kept their stale cache. Now the whole sheet is recomputed at save."""
    store = _seeded_store()
    tools = {t.name: t for t in create_table_tools(store, evaluator=_GRID_EVALUATOR)}
    reply = tools["write_table_cells"].func(
        path="book.table.json", sheet="s0", cells={"J1": "=ROUND(H1*I1*2,0)"},
        description="amount", revision=store.read("t1", "book.table.json").revision,
        runtime=_runtime(thread_id="t1"),
    )
    assert "Formulas: J1 =ROUND(H1*I1*2,0) → 42" in reply
    assert "1 other formula cell(s) recomputed with them" in reply  # the stale A6
    saved = json.loads(store.read("t1", "book.table.json").content)
    cells = {(c["r"], c["c"]): c["v"] for c in saved["data"]["sheet"][0]["celldata"]}
    written = cells[(0, 9)]
    assert written["f"] == "=ROUND(H1*I1*2,0)" and written["v"] == 42 and written["m"] == "42"
    assert written["ct"] == {"fa": "General", "t": "n"}
    dependent = cells[(5, 0)]
    assert dependent["v"] == 42 and dependent["m"] == "42"  # stale 999 refreshed


def test_a_formula_the_grid_cannot_run_is_flagged_not_stamped() -> None:
    store = _seeded_store()
    tools = {t.name: t for t in create_table_tools(store, evaluator=_GRID_EVALUATOR)}
    reply = tools["write_table_cells"].func(
        path="book.table.json", sheet="s0", cells={"J2": "=XLOOKUP(1,A1:A2,B1:B2)"},
        description="lookup", revision=store.read("t1", "book.table.json").revision,
        runtime=_runtime(thread_id="t1"),
    )
    assert "J2 =XLOOKUP(1,A1:A2,B1:B2) → #ERR" in reply
    saved = json.loads(store.read("t1", "book.table.json").content)
    cells = {(c["r"], c["c"]): c["v"] for c in saved["data"]["sheet"][0]["celldata"]}
    assert "v" not in cells[(1, 9)]  # never stamp a value the engine could not compute


def test_without_an_evaluator_the_reply_says_values_were_not_computed() -> None:
    store = _seeded_store()
    tools = {t.name: t for t in create_table_tools(store)}
    reply = tools["write_table_cells"].func(
        path="book.table.json", sheet="s0", cells={"J1": "=SUM(H1:I1)"},
        description="sum", revision=store.read("t1", "book.table.json").revision,
        runtime=_runtime(thread_id="t1"),
    )
    assert "not computed by the grid engine" in reply and "no evaluator configured" in reply


def test_without_an_evaluator_the_workbook_engine_takes_the_whole_save() -> None:
    """A deployment that ships LibreOffice but not the formula CLI still gets
    values on screen: with no light engine at all, every formula save falls
    through to the workbook engine instead of leaving blank cells."""
    from openpyxl import Workbook, load_workbook

    def fake_recalc(data: bytes) -> bytes:
        source = load_workbook(io.BytesIO(data))
        out = Workbook()
        out.remove(out.active)
        for ws in source.worksheets:
            fresh = out.create_sheet(ws.title)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
                    fresh.cell(row=cell.row, column=cell.column,
                               value=888 if is_formula else cell.value)
        buffer = io.BytesIO()
        out.save(buffer)
        return buffer.getvalue()

    store = _seeded_store()
    tools = {t.name: t for t in create_table_tools(store, xlsx_recalc=fake_recalc)}
    reply = tools["write_table_cells"].func(
        path="book.table.json", sheet="s0", cells={"J1": "=SUM(H1:I1)"},
        description="sum", revision=store.read("t1", "book.table.json").revision,
        runtime=_runtime(thread_id="t1"),
    )
    assert "Recalculated with the full spreadsheet engine" in reply
    saved = json.loads(store.read("t1", "book.table.json").content)
    cells = {(c["r"], c["c"]): c["v"] for c in saved["data"]["sheet"][0]["celldata"]}
    written = cells[(0, 9)]
    assert written["v"] == 888 and written["f"] == "=SUM(H1:I1)"


def test_a_formula_the_grid_cannot_run_falls_through_to_the_workbook_engine() -> None:
    """The light engine flags #ERR; the host's xlsx_recalc (LibreOffice in the
    reference deployment) recalculates the exported workbook and its values
    land on the cells — the engine that opens the file fills the screen."""
    from openpyxl import Workbook, load_workbook

    def fake_recalc(data: bytes) -> bytes:
        # Stands in for the LibreOffice endpoint: same sheet shape, every
        # formula cell answered with 777.
        source = load_workbook(io.BytesIO(data))
        out = Workbook()
        out.remove(out.active)
        for ws in source.worksheets:
            fresh = out.create_sheet(ws.title)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
                    fresh.cell(row=cell.row, column=cell.column,
                               value=777 if is_formula else cell.value)
        buffer = io.BytesIO()
        out.save(buffer)
        return buffer.getvalue()

    store = _seeded_store()
    tools = {t.name: t for t in create_table_tools(
        store, evaluator=_GRID_EVALUATOR, xlsx_recalc=fake_recalc)}
    reply = tools["write_table_cells"].func(
        path="book.table.json", sheet="s0", cells={"J2": "=XLOOKUP(1,A1:A2,B1:B2)"},
        description="lookup", revision=store.read("t1", "book.table.json").revision,
        runtime=_runtime(thread_id="t1"),
    )
    assert "Recalculated with the full spreadsheet engine: J2 → 777" in reply
    saved = json.loads(store.read("t1", "book.table.json").content)
    cells = {(c["r"], c["c"]): c["v"] for c in saved["data"]["sheet"][0]["celldata"]}
    lookup = cells[(1, 9)]
    assert lookup["v"] == 777 and lookup["m"] == "777"
    assert lookup["f"] == "=XLOOKUP(1,A1:A2,B1:B2)"  # still a live formula


def test_a_failing_workbook_engine_leaves_an_honest_note_and_the_save_lands() -> None:
    def broken_recalc(data: bytes) -> bytes:
        raise RuntimeError("endpoint unreachable")

    store = _seeded_store()
    tools = {t.name: t for t in create_table_tools(
        store, evaluator=_GRID_EVALUATOR, xlsx_recalc=broken_recalc)}
    reply = tools["write_table_cells"].func(
        path="book.table.json", sheet="s0", cells={"J2": "=XLOOKUP(1,A1:A2,B1:B2)"},
        description="lookup", revision=store.read("t1", "book.table.json").revision,
        runtime=_runtime(thread_id="t1"),
    )
    assert "Full recalculation failed" in reply and "endpoint unreachable" in reply
    assert "revision v2" in reply  # the save landed anyway


def test_a_supported_formula_never_calls_the_workbook_engine() -> None:
    calls: list[int] = []

    def counting_recalc(data: bytes) -> bytes:
        calls.append(1)
        return data

    store = _seeded_store()
    tools = {t.name: t for t in create_table_tools(
        store, evaluator=_GRID_EVALUATOR, xlsx_recalc=counting_recalc)}
    tools["write_table_cells"].func(
        path="book.table.json", sheet="s0", cells={"J1": "=SUM(H1:I1)"},
        description="sum", revision=store.read("t1", "book.table.json").revision,
        runtime=_runtime(thread_id="t1"),
    )
    assert calls == []  # the light engine handled it; no seconds spent


import shutil as _shutil  # noqa: E402 — soffice discovery for the real-engine test
from pathlib import Path as _Path  # noqa: E402

_SOFFICE = _shutil.which("soffice") or (
    "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    if _Path("/Applications/LibreOffice.app/Contents/MacOS/soffice").exists()
    else None
)


@pytest.mark.skipif(_SOFFICE is None, reason="needs LibreOffice for the real recalc")
def test_libreoffice_really_fills_the_cells_the_grid_engine_cannot() -> None:
    """End to end with the real workbook engine: SUMPRODUCT — outside the
    grid engine's surface — lands on the cell with LibreOffice's value."""
    import subprocess
    import tempfile

    def soffice_recalc(data: bytes) -> bytes:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = _Path(tmp)
            (tmp_path / "in.xlsx").write_bytes(data)
            profile = tmp_path / "profile" / "user"
            profile.mkdir(parents=True)
            (profile / "registrymodifications.xcu").write_text(
                '<?xml version="1.0" encoding="UTF-8"?>\n'
                '<oor:items xmlns:oor="http://openoffice.org/2001/registry">\n'
                ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">\n'
                '  <prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop>\n'
                " </item>\n</oor:items>\n"
            )
            out = tmp_path / "out"
            out.mkdir()
            subprocess.run(
                [_SOFFICE, "--headless", "--norestore", "--nolockcheck", "--nodefault",
                 f"-env:UserInstallation={(tmp_path / 'profile').as_uri()}",
                 "--convert-to", "xlsx", "--outdir", str(out), str(tmp_path / "in.xlsx")],
                check=True, capture_output=True, timeout=120,
            )
            return (out / "in.xlsx").read_bytes()

    store = _seeded_store()
    tools = {t.name: t for t in create_table_tools(
        store, evaluator=_GRID_EVALUATOR, xlsx_recalc=soffice_recalc)}
    reply = tools["write_table_cells"].func(
        path="book.table.json", sheet="s0",
        cells={"J2": "=XLOOKUP(1,A1:A2,B1:B2)", "J3": "=SUMPRODUCT(H1:I1,H1:I1)"},
        description="beyond the grid engine",
        revision=store.read("t1", "book.table.json").revision,
        runtime=_runtime(thread_id="t1"),
    )
    assert "Recalculated with the full spreadsheet engine" in reply
    saved = json.loads(store.read("t1", "book.table.json").content)
    cells = {(c["r"], c["c"]): c["v"] for c in saved["data"]["sheet"][0]["celldata"]}
    assert cells[(2, 9)]["v"] == 5 * 5 + 14300 * 14300  # SUMPRODUCT, LibreOffice's number
    assert cells[(2, 9)]["f"] == "=SUMPRODUCT(H1:I1,H1:I1)"
