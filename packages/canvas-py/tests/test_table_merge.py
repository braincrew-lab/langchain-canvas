"""rows <-> sheet convergence — mirror of canvas-react's tableMerge tests.

The four design rules pinned here: (1) a typed formula projects as its
source text, never its cached value; (2) agent row deletion truncates
trailing sheet rows, formatting included; (3) cells outside the rows
rectangle are preserved (a documented limit — they stay invisible to rows);
(4) type-normalized comparison, so 80 vs "80" never reads as a change.
"""

from __future__ import annotations

import io
import json

from openpyxl import load_workbook

from langchain_canvas import encode_artifact
from langchain_canvas.exporters import TableXlsxExporter
from langchain_canvas.table_merge import (
    merge_rows_into_sheet,
    project_sheet_into_rows,
    same_cell_content,
)

COLS = [{"key": "dept", "label": "Dept"}, {"key": "amount", "label": "Amount"}]
HEADER = [
    {"r": 0, "c": 0, "v": {"v": "Dept", "m": "Dept", "bl": 1}},
    {"r": 0, "c": 1, "v": {"v": "Amount", "m": "Amount", "bl": 1}},
]


def _sheet(celldata: list[dict]) -> list[dict]:
    return [{"name": "Sheet1", "id": "s1", "row": 60, "column": 8, "celldata": celldata}]


def _cell(sheet: list[dict] | None, r: int, c: int) -> dict | None:
    assert sheet is not None
    for cell in sheet[0]["celldata"]:
        if cell["r"] == r and cell["c"] == c:
            return cell["v"]
    return None


def test_same_cell_content_normalizes_types():
    assert same_cell_content(80, {"v": "80", "m": "80"})
    assert same_cell_content("80", {"v": 80})
    assert same_cell_content("", None)
    assert same_cell_content("=SUM(B2:B4)", {"f": "SUM(B2:B4)", "v": 60})
    assert not same_cell_content(60, {"f": "=SUM(B2:B4)", "v": 60})


def test_project_prefers_formula_source_and_extends_rows():
    sheet = _sheet(
        [
            *HEADER,
            {"r": 1, "c": 1, "v": {"f": "=SUM(B2:B2)", "v": 60, "m": "60"}},
            {"r": 2, "c": 0, "v": {"v": "added"}},
        ]
    )
    rows = project_sheet_into_rows(COLS, [{"dept": "x", "amount": 0}], sheet)
    assert rows[0]["amount"] == "=SUM(B2:B2)"
    assert rows[1]["dept"] == "added"


def test_project_drops_trailing_deleted_rows():
    sheet = _sheet([*HEADER, {"r": 1, "c": 0, "v": {"v": "A"}}])
    rows = project_sheet_into_rows(COLS, [{"dept": "A"}, {"dept": "gone"}], sheet)
    assert len(rows) == 1


def test_merge_is_noop_when_content_agrees():
    sheet = _sheet([*HEADER, {"r": 1, "c": 1, "v": {"v": "80", "m": "80"}}])
    assert merge_rows_into_sheet(COLS, [{"dept": "", "amount": 80}], sheet) is sheet


def test_merge_overrides_value_and_keeps_styling():
    sheet = _sheet([*HEADER, {"r": 1, "c": 1, "v": {"v": 80, "m": "80", "bl": 1, "bg": "#ff0"}}])
    merged = merge_rows_into_sheet(COLS, [{"dept": "", "amount": 95}], sheet)
    cell = _cell(merged, 1, 1)
    assert cell is not None
    assert cell["v"] == 95 and cell["bl"] == 1 and cell["bg"] == "#ff0"


def test_merge_appends_rows_truncates_ghosts_preserves_margin_notes():
    sheet = _sheet(
        [
            *HEADER,
            {"r": 1, "c": 0, "v": {"v": "keep"}},
            {"r": 2, "c": 0, "v": {"v": "ghost", "bl": 1}},
            {"r": 5, "c": 4, "v": {"v": "margin note"}},
        ]
    )
    merged = merge_rows_into_sheet(
        COLS, [{"dept": "keep"}, {"dept": "Total", "amount": "=SUM(B2:B2)"}], sheet
    )
    assert _cell(merged, 2, 0)["v"] == "Total"  # override beats the ghost slot
    total = _cell(merged, 2, 1)
    assert total is not None
    assert total["f"] == "=SUM(B2:B2)"
    assert _cell(merged, 5, 4)["v"] == "margin note"


def test_encode_artifact_projects_sheet_into_rows_on_save():
    artifact = {
        "type": "table",
        "title": "T",
        "data": {
            "columns": COLS,
            "rows": [{"dept": "Sales", "amount": 60}],
            "sheet": _sheet(
                [
                    *HEADER,
                    {"r": 1, "c": 0, "v": {"v": "Sales"}},
                    {"r": 1, "c": 1, "v": {"v": 75, "m": "75"}},  # person: 60 -> 75
                ]
            ),
        },
    }
    stored = json.loads(encode_artifact(artifact, "t.table.json"))
    assert stored["data"]["rows"] == [{"dept": "Sales", "amount": 75}]


def test_xlsx_export_includes_agent_rows_and_person_styling():
    content = json.dumps(
        {
            "type": "table",
            "data": {
                "columns": COLS,
                "rows": [
                    {"dept": "Sales", "amount": 95},  # agent changed 80 -> 95
                    {"dept": "Total", "amount": "=SUM(B2:B2)"},  # agent-added row
                ],
                "sheet": _sheet(
                    [
                        *HEADER,
                        {"r": 1, "c": 0, "v": {"v": "Sales"}},
                        {"r": 1, "c": 1, "v": {"v": 80, "m": "80", "bl": 1}},
                    ]
                ),
            },
        }
    )
    result = TableXlsxExporter().export(content, path="t.table.json")
    ws = load_workbook(io.BytesIO(result.data))["Sheet1"]
    assert ws.cell(row=2, column=2).value == 95  # agent value wins its cell
    assert ws.cell(row=3, column=2).value == "=SUM(B2:B2)"  # agent row exported
