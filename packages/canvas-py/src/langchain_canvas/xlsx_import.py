"""Rich ``.xlsx`` -> Fortune-sheet import — the twin of
``canvas-react/src/io/xlsx.ts``.

The browser reads an uploaded workbook with exceljs; this reads the same
bytes with openpyxl and must land on the same sheets, or a table a person
dragged in and a table an agent built stop looking alike.
``tests/test_xlsx_parity.py`` holds both to one golden file.

The plain ``{columns, rows}`` shape throws away everything that makes a real
spreadsheet look like one — fonts, colours, fills, number formats, merged
cells, column widths, and every sheet after the first. This builds Fortune's
native sheet model instead and *also* returns a flat columns/rows view of the
first sheet for export and fallback.
"""

from __future__ import annotations

import base64
import io
import math
import re
from datetime import date, datetime, time
from typing import Any

#: Excel's default (Office) theme palette, indexed as the file reports it.
#: Real spreadsheets colour most cells by theme index + tint rather than a
#: literal ARGB, so resolving these is what brings a file's colours across.
THEME_PALETTE = [
    "FFFFFF", "000000", "E7E6E6", "44546A",  # lt1, dk1, lt2, dk2
    "4472C4", "ED7D31", "A5A5A5", "FFC000",  # accent 1-4
    "5B9BD5", "70AD47", "0563C1", "954F72",  # accent 5-6, hlink, followed-hlink
]

# Legacy indexed colour palette (BIFF8). Many real files still colour cells by
# a palette index rather than ARGB/theme, so without this those colours vanish.
INDEXED_PALETTE: dict[int, str] = {
    0: "000000", 1: "FFFFFF", 2: "FF0000", 3: "00FF00", 4: "0000FF", 5: "FFFF00",
    6: "FF00FF", 7: "00FFFF", 8: "000000", 9: "FFFFFF", 10: "FF0000", 11: "00FF00",
    12: "0000FF", 13: "FFFF00", 14: "FF00FF", 15: "00FFFF", 16: "800000", 17: "008000",
    18: "000080", 19: "808000", 20: "800080", 21: "008080", 22: "C0C0C0", 23: "808080",
    24: "9999FF", 25: "993366", 26: "FFFFCC", 27: "CCFFFF", 28: "660066", 29: "FF8080",
    30: "0066CC", 31: "CCCCFF", 32: "000080", 33: "FF00FF", 34: "FFFF00", 35: "00FFFF",
    36: "800080", 37: "800000", 38: "008080", 39: "0000FF", 40: "00CCFF", 41: "CCFFFF",
    42: "CCFFCC", 43: "FFFF99", 44: "99CCFF", 45: "FF99CC", 46: "CC99FF", 47: "FFCC99",
    48: "3366FF", 49: "33CCCC", 50: "99CC00", 51: "FFCC00", 52: "FF9900", 53: "FF6600",
    54: "666699", 55: "969696", 56: "003366", 57: "339966", 58: "003300", 59: "333300",
    60: "993300", 61: "993366", 62: "333399", 63: "333333",
}

_H_ALIGN = {"center": 0, "left": 1, "right": 2}
# openpyxl says "center" where exceljs says "middle" for the vertical axis.
_V_ALIGN = {"center": 0, "middle": 0, "top": 1, "bottom": 2}

#: Border style name -> Fortune-sheet border style number.
_BORDER_STYLE = {
    "hair": 2, "thin": 1, "dotted": 3, "dashDot": 5, "dashDotDot": 6, "dashed": 4,
    "mediumDashed": 9, "mediumDashDot": 10, "mediumDashDotDot": 11, "slantDashDot": 12,
    "medium": 8, "double": 7, "thick": 13,
}

_EMU_PER_PX = 9525
_DEFAULT_COL_PX = 73  # Fortune's default column width / row height, used to
_DEFAULT_ROW_PX = 19  # place an image anchored to a cell with no explicit size

_MONTHS_LONG = ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]
_MONTHS_SHORT = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                 "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
_DATE_TOKENS = re.compile(r"yyyy|yy|mmmm|mmm|mm|m|dd|d|hh|h|ss")
_ESCAPED = re.compile(r"\\(.)")


def _number_format(cell: Any) -> str:
    """A cell's number format with its escapes taken out.

    A file writes ``yyyy\\-mm\\-dd`` and ``\\$#,##0``; the backslash marks the
    next character as literal. The browser's reader drops those marks and
    openpyxl keeps them, so drop them here or every escaped format reads
    differently on the two sides.
    """
    return _ESCAPED.sub(r"\1", cell.number_format or "")


def _round(value: float) -> int:
    """Round half away from zero the way the browser does — Python's own
    ``round`` breaks ties to even, which puts the two twins one pixel apart."""
    return math.floor(value + 0.5)


def _tint_channel(channel: int, tint: float) -> int:
    """Apply an Excel tint (-1..1) to one 0-255 channel: <0 darkens, >0 lightens."""
    value = channel * (1 + tint) if tint < 0 else channel * (1 - tint) + 255 * tint
    return max(0, min(255, _round(value)))


def _to_hex(color: Any) -> str | None:
    """A colour (ARGB, theme index + tint, or legacy indexed) as CSS ``#RRGGBB``."""
    if color is None:
        return None
    kind = getattr(color, "type", None)
    hex_value: str | None = None
    if kind == "rgb":
        raw = getattr(color, "rgb", None)
        if isinstance(raw, str):
            hex_value = raw[2:] if len(raw) == 8 else raw
    elif kind == "theme":
        index = getattr(color, "theme", None)
        if isinstance(index, int) and 0 <= index < len(THEME_PALETTE):
            hex_value = THEME_PALETTE[index]
    elif kind == "indexed":
        # 64/65 (system auto) intentionally absent
        hex_value = INDEXED_PALETTE.get(getattr(color, "indexed", -1))
    if not hex_value or not re.fullmatch(r"[0-9a-fA-F]{6}", hex_value):
        return None
    tint = getattr(color, "tint", 0) or 0
    if tint:
        number = int(hex_value, 16)
        red = _tint_channel((number >> 16) & 255, tint)
        green = _tint_channel((number >> 8) & 255, tint)
        blue = _tint_channel(number & 255, tint)
        hex_value = f"{(red << 16) | (green << 8) | blue:06x}"
    return f"#{hex_value}"


def _border_side(side: Any) -> dict[str, Any] | None:
    """One border side in Fortune's shape, or ``None`` if the cell has none."""
    style = getattr(side, "style", None)
    if not style:
        return None
    return {
        "style": _BORDER_STYLE.get(style, 1),
        "color": _to_hex(getattr(side, "color", None)) or "#000000",
    }


def _number(value: float) -> Any:
    """A float that is whole reads as an integer, the way JavaScript prints it."""
    return int(value) if isinstance(value, float) and value.is_integer() else value


def _display(value: Any) -> str:
    """Display text for a value."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        return _format_date(value, "yyyy-mm-dd")
    if isinstance(value, date):
        return _format_date(datetime(value.year, value.month, value.day), "yyyy-mm-dd")
    if isinstance(value, float):
        return _js_number(value)
    return str(value)


def _js_number(value: float) -> str:
    """A float printed the way JavaScript prints it: no trailing ``.0``."""
    if value != value or value in (float("inf"), float("-inf")):
        return str(value)
    if float(value).is_integer():
        return str(int(value))
    return repr(float(value))


def _format_number(value: float, num_fmt: str | None) -> str:
    """Render a number the way its Excel number format would — thousands
    separators, fixed decimals, percent, and a leading currency symbol. Not a
    full format engine, but it covers the everyday patterns so cells read like
    the source ("1,234.50", "15.6%", "$1,000") instead of a bare "1234.5"."""
    fmt = num_fmt if num_fmt and num_fmt != "General" else ""
    if not fmt:
        return _js_number(value)
    found = re.search(r"\.([0#]+)", fmt)
    decimals = len(found.group(1)) if found else 0
    if "%" in fmt:
        return f"{value * 100:.{decimals}f}%"
    thousands = re.search(r"[#0],[#0]", fmt) is not None
    out = f"{value:,.{decimals}f}" if thousands else f"{value:.{decimals}f}"
    currency = re.search(r"[$₩€£¥]", fmt)
    if currency:
        symbol = currency.group(0)
        out = f"-{symbol}{out[1:]}" if value < 0 else f"{symbol}{out}"
    return out


def _format_date(moment: datetime, num_fmt: str | None) -> str:
    """Format a datetime by an Excel date pattern (yyyy/yy, mm/m, dd/d, hh/h,
    ss). Non-token text (Korean 년/월/일 and the like) passes through."""
    dated = bool(num_fmt) and num_fmt != "General" and re.search(r"[ymdhs]", num_fmt or "", re.I)
    fmt = num_fmt if dated and num_fmt else "yyyy-mm-dd"
    table = {
        "yyyy": str(moment.year), "yy": f"{moment.year % 100:02d}",
        "mmmm": _MONTHS_LONG[moment.month - 1], "mmm": _MONTHS_SHORT[moment.month - 1],
        "mm": f"{moment.month:02d}", "m": str(moment.month),
        "dd": f"{moment.day:02d}", "d": str(moment.day),
        "hh": f"{moment.hour:02d}", "h": str(moment.hour),
        "ss": f"{moment.second:02d}",
    }
    # Longest tokens first so "yyyy" wins over "yy", "mmmm" over "mm".
    return _DATE_TOKENS.sub(lambda t: table.get(t.group(0), t.group(0)), fmt)


def _styled(cell: Any, default_style: Any) -> bool:
    """Whether the cell names a format of its own.

    exceljs reports no font, fill, alignment or number format for a cell that
    points at the workbook's default format; openpyxl resolves the default for
    every cell instead, so ask the question directly. Without this the twin
    would dress every plain cell in the default font and drift from the
    browser on the very first file.
    """
    style = getattr(cell, "_style", None)
    return style is not None and style != default_style


def _cell_value(cell: Any, cached: Any, styled: bool) -> dict[str, Any] | None:
    """A Fortune cell ``v`` object from one cell, carrying its style.
    Merges are handled by the caller, so this reads only the cell's own value."""
    raw = cell.value
    is_formula = isinstance(raw, str) and raw.startswith("=")
    result = cached if is_formula else raw
    num_fmt = _number_format(cell) if styled else None
    text = _display(result)
    if text == "" and not is_formula and not styled:
        return None

    v: dict[str, Any] = {"m": text}
    if is_formula:
        v["f"] = raw
        v["v"] = _display(result)
        if isinstance(result, (int, float)) and not isinstance(result, bool):
            v["m"] = _format_number(float(result), num_fmt)
    elif isinstance(result, (int, float)) and not isinstance(result, bool):
        v["v"] = _number(result)
        # show "1,234.50"/"15.6%"/"$1,000", not "1234.5"
        v["m"] = _format_number(float(result), num_fmt)
        v["ct"] = {"fa": num_fmt or "General", "t": "n"}
    elif isinstance(result, (datetime, date)) and not isinstance(result, time):
        moment = (result if isinstance(result, datetime)
                  else datetime(result.year, result.month, result.day))
        v["m"] = _format_date(moment, num_fmt)
        v["v"] = v["m"]
        v["ct"] = {"fa": num_fmt or "yyyy-mm-dd", "t": "d"}
    else:
        v["v"] = text

    if not styled:
        return v

    font = cell.font
    if font.b:
        v["bl"] = 1
    if font.i:
        v["it"] = 1
    if font.u:
        v["un"] = 1
    if font.sz:
        v["fs"] = _number(font.sz)
    if font.name:
        v["ff"] = font.name
    font_colour = _to_hex(font.color)
    if font_colour:
        v["fc"] = font_colour

    # Solid fill -> background colour. The fill colour lives in `fgColor`; some
    # writers only populate `bgColor`, so fall back to it.
    fill = cell.fill
    if getattr(fill, "patternType", None) == "solid":
        background = _to_hex(fill.fgColor) or _to_hex(fill.bgColor)
        if background:
            v["bg"] = background
    alignment = cell.alignment
    if alignment.horizontal in _H_ALIGN:
        v["ht"] = _H_ALIGN[alignment.horizontal]
    if alignment.vertical in _V_ALIGN:
        v["vt"] = _V_ALIGN[alignment.vertical]
    if alignment.wrap_text:
        v["tb"] = "2"  # Fortune: 2 = wrap text
    return v


def _parse_merge(merged: Any) -> tuple[str, dict[str, int]]:
    """A merge range as Fortune's ``{r,c,rs,cs}`` (0-based) plus its key."""
    row, column = merged.min_row - 1, merged.min_col - 1
    return (
        f"{row}_{column}",
        {"r": row, "c": column,
         "rs": merged.max_row - merged.min_row + 1,
         "cs": merged.max_col - merged.min_col + 1},
    )


def _sheet_images(
    worksheet: Any, columnlen: dict[int, int], rowlen: dict[int, int]
) -> list[dict[str, Any]]:
    """A worksheet's floating images as Fortune images (px position + data URL)."""
    placed = list(getattr(worksheet, "_images", []) or [])
    if not placed:
        return []

    def col_left(column: int) -> int:
        return sum(columnlen.get(i, _DEFAULT_COL_PX) for i in range(column))

    def row_top(row: int) -> int:
        return sum(rowlen.get(i, _DEFAULT_ROW_PX) for i in range(row))

    images: list[dict[str, Any]] = []
    for index, image in enumerate(placed):
        anchor = getattr(image, "anchor", None)
        start = getattr(anchor, "_from", None)
        if start is None:
            continue
        try:
            blob = image._data()
        except Exception:  # noqa: BLE001 — an unreadable picture is not a reason to lose the sheet
            continue
        kind = (getattr(image, "format", None) or "png").lower()
        src = f"data:image/{kind};base64,{base64.b64encode(blob).decode('ascii')}"
        left = col_left(start.col) + _round((start.colOff or 0) / _EMU_PER_PX)
        top = row_top(start.row) + _round((start.rowOff or 0) / _EMU_PER_PX)
        extent = getattr(anchor, "ext", None)
        width = _round(extent.cx / _EMU_PER_PX) if extent else 0
        height = _round(extent.cy / _EMU_PER_PX) if extent else 0
        end = getattr(anchor, "to", None)
        if (not width or not height) and end is not None:
            # A two-cell anchor sizes the image by the span from start to end.
            width = col_left(end.col) - left
            height = row_top(end.row) - top
        images.append({
            "id": f"img_{index}",
            "src": src,
            "left": max(0, _round(left)),
            "top": max(0, _round(top)),
            "width": max(16, _round(width or 100)),
            "height": max(16, _round(height or 100)),
        })
    return images


def _column_letter(index: int) -> str:
    letters = ""
    while index > 0:
        letters = chr(65 + (index - 1) % 26) + letters
        index = (index - 1) // 26
    return letters


def _flat_value(value: Any) -> Any:
    """A cell's value with its type preserved (numbers stay numbers)."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return _number(value)
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _flatten(
    worksheet: Any, cached: Any, covered: dict[tuple[int, int], tuple[int, int]]
) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    """First-sheet columns/rows (values only, types preserved), read through merges."""
    if worksheet is None:
        return [], []
    height, width = worksheet.max_row, worksheet.max_column

    def read(row: int, column: int) -> Any:
        cell = worksheet.cell(row=row, column=column)
        value = cell.value
        if isinstance(value, str) and value.startswith("="):
            value = cached.cell(row=row, column=column).value
        if value is None:
            master = covered.get((row - 1, column - 1))
            if master is not None:
                value = worksheet.cell(row=master[0] + 1, column=master[1] + 1).value
        return _flat_value(value)

    seen: dict[str, int] = {}
    columns: list[dict[str, str]] = []
    for column in range(1, width + 1):
        label = str(read(1, column)).strip() or _column_letter(column)
        count = seen.get(label, 0)
        seen[label] = count + 1
        columns.append({"key": f"{label} ({count + 1})" if count else label, "label": label})

    rows: list[dict[str, Any]] = []
    for row in range(2, height + 1):
        entry: dict[str, Any] = {}
        has_value = False
        for column in range(1, width + 1):
            value = read(row, column)
            entry[columns[column - 1]["key"]] = value
            if value != "":
                has_value = True
        if has_value:
            rows.append(entry)
    return columns, rows


def xlsx_to_sheets(data: bytes) -> dict[str, Any]:
    """An .xlsx workbook as Fortune sheets (rich) plus a flat first-sheet view.

    Returns ``{"sheets": [...], "columns": [...], "rows": [...]}`` — the same
    shape ``xlsxToSheets`` returns in the browser. Raises
    :class:`~langchain_canvas.converters.MissingConverterDependencyError` when
    ``openpyxl`` is absent.
    """
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:
        from .converters import MissingConverterDependencyError

        raise MissingConverterDependencyError(
            "reading .xlsx needs openpyxl — install langchain-canvas[xlsx] "
            "or register your own converter for .xlsx"
        ) from exc

    # Two passes over the same bytes: one keeps the typed formulas, the other
    # the results Excel cached for them. openpyxl gives one or the other.
    workbook = load_workbook(io.BytesIO(data))
    values = load_workbook(io.BytesIO(data), data_only=True)
    default_style = workbook._cell_styles[0] if workbook._cell_styles else None

    sheets: list[dict[str, Any]] = []
    for index, worksheet in enumerate(workbook.worksheets):
        cached = values.worksheets[index]
        height, width = worksheet.max_row, worksheet.max_column

        # Merges first — Fortune needs both a `config.merge` entry and per-cell
        # `mc` markers: the master carries the value + span, the covered cells
        # only point back to it, so the content shows once across the span.
        merge: dict[str, dict[str, int]] = {}
        covered: dict[tuple[int, int], tuple[int, int]] = {}
        for merged in worksheet.merged_cells.ranges:
            key, entry = _parse_merge(merged)
            merge[key] = entry
            for row in range(entry["r"], entry["r"] + entry["rs"]):
                for column in range(entry["c"], entry["c"] + entry["cs"]):
                    if (row, column) != (entry["r"], entry["c"]):
                        covered[(row, column)] = (entry["r"], entry["c"])

        celldata: list[dict[str, Any]] = []
        border_info: list[dict[str, Any]] = []
        for row in range(1, height + 1):
            for column in range(1, width + 1):
                cell = worksheet.cell(row=row, column=column)
                styled = _styled(cell, default_style)
                # Borders are captured for every cell (merged ones too — a
                # box's edges live on its outer cells) so the grid's rules
                # match the source exactly.
                if styled:
                    border = cell.border
                    sides = {
                        "l": _border_side(border.left), "r": _border_side(border.right),
                        "t": _border_side(border.top), "b": _border_side(border.bottom),
                    }
                    if any(sides.values()):
                        value: dict[str, Any] = {"row_index": row - 1, "col_index": column - 1}
                        value.update({k: side for k, side in sides.items() if side})
                        border_info.append({"rangeType": "cell", "value": value})

                at = (row - 1, column - 1)
                master = covered.get(at)
                if master is not None:
                    celldata.append({"r": at[0], "c": at[1],
                                     "v": {"mc": {"r": master[0], "c": master[1]}}})
                    continue
                v = _cell_value(cell, cached.cell(row=row, column=column).value, styled)
                if v is not None:
                    span = merge.get(f"{at[0]}_{at[1]}")
                    if span:
                        v["mc"] = span  # this cell is a merge master
                    celldata.append({"r": at[0], "c": at[1], "v": v})

        # Auto-fit each column to its widest cell so short content isn't padded
        # out to wide empty columns. A stored column width is honoured as a
        # floor. Merged cells span several columns, so they size none of them.
        col_chars: dict[int, int] = {}
        for cell_entry in celldata:
            v = cell_entry["v"]
            if "mc" in v:
                continue
            text = v.get("m")
            length = len(text) if isinstance(text, str) else 0
            column = cell_entry["c"]
            if length > col_chars.get(column, 0):
                col_chars[column] = length
        columnlen: dict[int, int] = {}
        for column in range(width):
            # `.get`, not `[]` — indexing this holder invents a column with the
            # default width, which would pad every unsized column to one size.
            dimension = worksheet.column_dimensions.get(_column_letter(column + 1))
            stored = dimension.width if dimension is not None else None
            pixels = max(
                col_chars[column] * 8 + 20 if col_chars.get(column) else 0,
                stored * 7 + 5 if stored else 0,
            )
            if pixels:
                columnlen[column] = max(56, min(320, _round(pixels)))
        rowlen: dict[int, int] = {}
        for row in range(1, height + 1):
            dimension = worksheet.row_dimensions.get(row)
            stored_height = dimension.height if dimension is not None else None
            if stored_height:
                rowlen[row - 1] = _round(stored_height * 1.33)  # points -> px

        sheets.append({
            "name": worksheet.title or f"Sheet{index + 1}",
            "id": f"sheet_{index}",
            "order": index,
            "status": 1 if index == 0 else 0,
            # Size the grid to the data plus a small buffer — enough to feel
            # like a real sheet, tight enough that the scrollbars stay
            # proportional and there is little empty grid to scroll past.
            "row": max(height + 8, 24),
            "column": max(width + 2, 10),
            "celldata": celldata,
            "config": {"merge": merge, "columnlen": columnlen,
                       "rowlen": rowlen, "borderInfo": border_info},
            "images": _sheet_images(worksheet, columnlen, rowlen),
        })

    first = workbook.worksheets[0] if workbook.worksheets else None
    first_covered: dict[tuple[int, int], tuple[int, int]] = {}
    if first is not None:
        for merged in first.merged_cells.ranges:
            _, entry = _parse_merge(merged)
            for row in range(entry["r"], entry["r"] + entry["rs"]):
                for column in range(entry["c"], entry["c"] + entry["cs"]):
                    if (row, column) != (entry["r"], entry["c"]):
                        first_covered[(row, column)] = (entry["r"], entry["c"])
    first_values = values.worksheets[0] if first is not None else None
    columns, rows = _flatten(first, first_values, first_covered)
    return {"sheets": sheets, "columns": columns, "rows": rows}
