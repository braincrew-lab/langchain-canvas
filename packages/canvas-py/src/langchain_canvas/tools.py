"""Standard canvas tools — the agent's hands on a persistent canvas.

Four file-level primitives over a :class:`~langchain_canvas.store.CanvasStore`:

- ``read_canvas``  — file content with line numbers + the current revision
- ``write_canvas`` — create or fully replace a file
- ``edit_canvas``  — replace one unique occurrence, **requires the revision
  returned by a prior read** (read-before-update enforced by the contract,
  not by prompt discipline)
- ``list_canvas_files`` — files currently on the canvas

The tools persist through the store **and** broadcast each committed
artifact-file change (``.html`` pages, ``.table.json`` tables) as wire events
(``canvas.create``/``patch``/``commit``) through the run's stream writer, so a
connected client redraws live. Without a stream
writer (unit tests, plain scripts) the broadcast is a silent no-op — same
contract as :class:`~langchain_canvas.emitter.Canvas`. Which canvas they act
on is resolved per call: ``canvas_id`` in the runtime context (or
``configurable``), falling back to ``thread_id`` — by default a thread and
its canvas are the same scope.

Build them with :func:`create_canvas_tools`, which closes over your store::

    store = InMemoryCanvasStore()
    agent = create_canvas_agent(model, tools=create_canvas_tools(store))
"""

from __future__ import annotations

import base64
import binascii
import hashlib
import io
import json
import re
import subprocess
from collections.abc import Callable, Sequence
from typing import Any

from langchain.tools import ToolRuntime, tool

from .assets import (
    ASSET_IMAGE_MIME,
    ASSETS_PREFIX,
    inline_canvas_assets,
    inline_slides_assets,
    normalize_asset_reference,
)
from .converters import (
    MAX_IMAGES_PER_CALL,
    MissingConverterDependencyError,
    PageRenderable,
    SourceConverter,
    UnsafeArchiveError,
    converter_for,
    default_converters,
    ensure_archive_within_limits,
)
from .deck_outline import deck_outline, deck_projection
from .document_lint import (
    format_document_warnings,
    is_document_path,
    lint_document_content,
)
from .document_ops import (
    DOCUMENT_OP_SUFFIXES,
    DocumentOpError,
    MissingDocumentDependencyError,
    insert_image,
    insert_paragraph,
    remove_paragraph,
    reopens,
    replace_image,
    replace_text,
)
from .exporters import (
    DEFAULT_SLIDE_PAGE_IN,
    ExportedFile,
    Exporter,
    MissingExporterDependencyError,
    TableXlsxExporter,
    default_exporters,
    exporter_for,
    pptx_page_size_inches,
)
from .formulas import SUPPORTED_FORMULA_FUNCTIONS, formula_guidance
from .layout_lint import (
    blocking_deck_findings,
    format_layout_warnings,
    lint_slides_data,
)
from .protocol.artifacts import TableData
from .replay import (
    ARTIFACT_SUFFIXES,
    SLIDES_SUFFIX,
    TABLE_SUFFIX,
    display_title,
    encode_artifact,
    encode_slides,
    events_for_commit,
    source_preview_events,
    working_copy_path,
)
from .state import last_change_line
from .store import (
    BinaryContentError,
    CanvasFileNotFoundError,
    CanvasStore,
    CanvasStoreError,
    Commit,
    EditConflictError,
    RevisionMismatchError,
)
from .table_outline import add_sheet as table_add_sheet
from .table_outline import table_view
from .table_outline import write_cells as table_write_cells

_RETRY_HINT = "Call read_canvas again and retry with the fresh revision and exact content."
_SOURCES_PREFIX = "sources/"
_SOURCES_READONLY = (
    "Error: files under sources/ are the user's original uploads and are "
    "read-only for the agent. Create a new canvas file instead (for example "
    "an .html page or a .table.json table)."
)
_SOURCES_READONLY_DOCUMENT = (
    "Error: {path} is the user's original upload and is read-only. Call "
    "open_document_for_editing on it first, then edit the copy it makes."
)
_SOURCES_READONLY_DECK = (
    "Error: {path} is the user's original upload and is read-only. Call "
    "open_deck_for_editing on it first, then edit the copy it makes."
)
_SOURCES_READONLY_WORKBOOK = (
    "Error: {path} is the user's original upload and is read-only. Its editable "
    'working copy is {copy} — read that with sheet="s0", then change cells with '
    "write_table_cells. Do not rebuild it with write_canvas."
)
_DEFAULT_READ_LIMIT = 400
#: Slides rendered alongside a save-time finding — enough to see the problem,
#: few enough that a deck with findings on every slide does not become a
#: slideshow in the tool result.
_EYE_MAX_SLIDES = 2


#: The width an eye image is resized to — enough to read a finding, a third
#: of the tokens of a full render on pixel-priced vision models.
_EYE_MAX_WIDTH = 1024


def _glance_size(block: dict) -> dict:
    """The image block resized to glance size (JPEG), or unchanged.

    A deck page arrived at 1921x1080 PNG, ~300KB and ~1.5k vision tokens,
    twice per save. The eye exists to show a finding, not to archive the
    slide; anything that fails here (no pillow, odd bytes) passes through.
    """
    try:
        import base64
        import io

        from PIL import Image  # type: ignore[import-untyped]

        raw = base64.b64decode(block.get("data", ""))
        image = Image.open(io.BytesIO(raw))
        if image.width <= _EYE_MAX_WIDTH:
            return block
        height = round(image.height * _EYE_MAX_WIDTH / image.width)
        resized = image.convert("RGB").resize((_EYE_MAX_WIDTH, height))
        out = io.BytesIO()
        resized.save(out, format="JPEG", quality=80)
        return {
            **block,
            "data": base64.b64encode(out.getvalue()).decode(),
            "mime_type": "image/jpeg",
        }
    except Exception:  # noqa: BLE001 - the eye is a bonus, never a failure
        return block


def _with_eye(text: str, images: list[dict]) -> str | list[dict]:
    """The tool reply as text, or text followed by page images when there are any."""
    if not images:
        return text
    return [{"type": "text", "text": text}, *images]
_DOCUMENT_FORMATS = ", ".join(DOCUMENT_OP_SUFFIXES)


def _is_document_file(path: str) -> bool:
    """Whether these are the binary documents the document operations edit."""
    return path.lower().endswith(DOCUMENT_OP_SUFFIXES)


def _other_ways_to_open(path: str, converters: list[SourceConverter]) -> str:
    """What to do with a file the document operations cannot open.

    A deck is the case that matters: there is no import from ``.pptx`` to a
    slides artifact yet, and an agent told only "this opens .docx" tends to
    invent a summary instead of saying so. Name the reads that do work —
    pages as images where a renderer covers the format, text otherwise — so
    the answer is a next step rather than a dead end.
    """
    if path.lower().endswith(".pptx"):
        return (
            f"To edit it, copy it out with open_deck_for_editing (it becomes a "
            f"{SLIDES_SUFFIX} deck you can edit and export back to PowerPoint). "
            f"To just look, read {path} — with `pages` for the slide images."
        )
    if any(
        isinstance(c, PageRenderable) and path.lower().endswith(c.suffixes)
        for c in converters
    ):
        return (
            f"Read {path} with `pages` to see it as images (or without `pages` "
            "for its text) — it stays a source file either way. Text canvas "
            "files are already editable: read one and use edit_canvas."
        )
    return (
        "Text canvas files are already editable — read one and use edit_canvas."
    )


_WORKING_COPY_MARKER = "Editing - "


# Markdown image syntax: `![alt](path)`. A Word file has no such shorthand and
# shows it as the literal characters, so a paragraph carrying one is refused
# and told where the picture goes instead.
_MARKDOWN_IMAGE = re.compile(r"!\[[^\]]*\]\([^)]*\)")

def _working_copy_name(source: str) -> str:
    """Canvas-root name for the editable copy of ``source``.

    The copy and the original are the same document, so nothing in the name
    itself tells them apart — the marker does. It goes in front because tabs
    show as much of a name as fits and clip the rest, and these names are
    long; a mark at the end would be the first thing to disappear. Copying a
    copy does not stack markers.
    """
    name = source.rsplit("/", 1)[-1]
    return name if name.startswith(_WORKING_COPY_MARKER) else _WORKING_COPY_MARKER + name


# An inline picture the deck reader hands over, by the four types it emits.
_IMAGE_DATA_URI = re.compile(r"^data:(image/(?:png|jpeg|gif|webp));base64,(.+)$", re.DOTALL)
_SUFFIX_FOR_MIME = {
    "image/png": ".png",
    "image/jpeg": ".jpg",
    "image/gif": ".gif",
    "image/webp": ".webp",
}


def _decode_image_data_uri(src: Any) -> tuple[str, bytes] | None:
    """(file suffix, bytes) for an inline image; ``None`` for anything else."""
    if not isinstance(src, str):
        return None
    match = _IMAGE_DATA_URI.match(src)
    if match is None:
        return None
    try:
        blob = base64.b64decode(match.group(2), validate=True)
    except (binascii.Error, ValueError):
        return None
    return _SUFFIX_FOR_MIME[match.group(1)], blob


def _deck_stem(target: str) -> str:
    """The deck's own name, without folder or the ``.slides.json`` suffix."""
    name = target.rsplit("/", 1)[-1]
    return name[: -len(SLIDES_SUFFIX)] if name.endswith(SLIDES_SUFFIX) else name


def _deck_copy_name(source: str) -> str:
    """Canvas-root name for the editable deck made from ``source``.

    The copy is a different kind of file from the upload — a deck the canvas
    owns, not a PowerPoint document — so it takes the suffix that says so
    rather than a marker on the same name.
    """
    name = source.rsplit("/", 1)[-1]
    stem = name[: -len(".pptx")] if name.lower().endswith(".pptx") else name
    return f"{stem}{SLIDES_SUFFIX}"


def _sources_readonly(path: str) -> str:
    """The refusal for an upload, naming the way forward for this file type.

    A Word file has one — copy it and edit the copy — and pointing at .html
    pages instead would be telling the agent it cannot do something it can.
    """
    if _is_document_file(path):
        return _SOURCES_READONLY_DOCUMENT.format(path=path)
    if path.lower().endswith(".pptx"):
        return _SOURCES_READONLY_DECK.format(path=path)
    if path.lower().endswith(".xlsx"):
        return _SOURCES_READONLY_WORKBOOK.format(path=path, copy=working_copy_path(path))
    return _SOURCES_READONLY


def _renderer_for(path: str, converters: list[SourceConverter]) -> PageRenderable | None:
    converter = converter_for(path, converters)
    return converter if isinstance(converter, PageRenderable) else None


def _verified(data: bytes, path: str, converters: list[SourceConverter]) -> str:
    """"" when the edited bytes pass every check, else why they did not.

    Two checks beyond the save-time one every file already gets: the result
    still opens as a document, and it still renders. A document that parses
    but no longer draws is exactly the failure a user finds by opening the
    file, and the caller refuses to save on either — an edit that cannot be
    seen is not an edit that gets reported as done.
    """
    problem = reopens(data)
    if problem is not None:
        return f"Error: the edit did not save — the result no longer opens ({problem})."
    renderer = _renderer_for(path, converters)
    if renderer is None:
        return ""
    try:
        renderer.render_pages(data, path=path, pages=[1])
    except Exception as exc:  # any renderer failure blocks the save
        return (
            f"Error: the edit did not save — the result no longer renders ({exc}). "
            "The file on the canvas is unchanged."
        )
    return ""


def _look_note(path: str, converters: list[SourceConverter]) -> str:
    """The instruction to actually look at what was just written."""
    if _renderer_for(path, converters) is None:
        return (
            " No page renderer is installed here, so the result was reopened but "
            "not seen — check it yourself before telling the user it is done."
        )
    return (
        f' Look at it before telling the user it is done: read_canvas(path="{path}", '
        'pages="grid").'
    )


def _broadcast_source(
    store: CanvasStore,
    runtime: ToolRuntime,
    canvas_id: str,
    path: str,
    is_new: bool,
    commit: Commit,
) -> None:
    """Push the file's card to a connected client (silent without a writer)."""
    writer = getattr(runtime, "stream_writer", None)
    if writer is None:
        return
    for event in source_preview_events(
        store,
        canvas_id,
        path,
        is_new=is_new,
        revision=commit.revision,
        description=commit.description,
    ):
        writer(event)


def _save_document(
    store: CanvasStore,
    runtime: ToolRuntime,
    canvas_id: str,
    path: str,
    data: bytes,
    description: str,
    revision: str,
    *,
    converters: list[SourceConverter],
    verb: str,
    note: str = "",
) -> str:
    """Check the edited bytes, save them, redraw the card, say what to look at.

    The order is the point: a result that no longer opens or no longer renders
    never reaches the store, so "saved" and "still a document" cannot come
    apart. The save itself keeps the read-before-write contract every other
    write keeps — a stale ``revision`` is refused, not overwritten.
    """
    problem = _verified(data, path, converters)
    if problem:
        return problem
    try:
        commit = store.write_bytes(
            canvas_id, path, data, description, base_revision=revision, actor="agent"
        )
    except RevisionMismatchError as exc:
        return f"Error: {exc}. {_RETRY_HINT}"
    except CanvasStoreError as exc:
        return f"Error: {exc}."
    _broadcast_source(store, runtime, canvas_id, path, False, commit)
    return f"{verb} {path} (revision {commit.revision}).{note}" + _look_note(path, converters)


def _canvas_id(runtime: ToolRuntime) -> str:
    """Resolve the target canvas for this call.

    Precedence: explicit ``canvas_id`` in the runtime context (attribute or
    mapping key), then ``configurable.canvas_id``, then ``configurable.thread_id``.
    """
    context = runtime.context
    for source in (
        getattr(context, "canvas_id", None),
        context.get("canvas_id") if isinstance(context, dict) else None,
    ):
        if source:
            return str(source)
    configurable: dict[str, Any] = (runtime.config or {}).get("configurable", {})
    for key in ("canvas_id", "thread_id"):
        if configurable.get(key):
            return str(configurable[key])
    raise ValueError(
        "No canvas id: provide `canvas_id` in the runtime context or run with a `thread_id`."
    )


def _sliced(content: str, offset: int, limit: int) -> tuple[str, str]:
    """A line window of ``content`` plus a continuation note ("" when complete)."""
    lines = content.split("\n")
    total = len(lines)
    window = lines[offset : offset + limit]
    numbered = "\n".join(
        f"{i:>4}\t{line}" for i, line in enumerate(window, start=offset + 1)
    )
    end = offset + len(window)
    if offset == 0 and end >= total:
        return numbered, ""
    note = f"[lines {offset + 1}-{end} of {total}"
    if end < total:
        note += f" — call read_canvas again with offset={end} for more"
    return numbered, note + "]"


def _parse_pages(spec: str) -> list[int]:
    """1-based page numbers from a spec like ``"3"``, ``"2-5"`` or ``"1,4,7"``.

    Order-preserving and de-duplicated; raises ``ValueError`` with an honest
    message on anything else, so the tool can relay it verbatim.
    """
    pages: list[int] = []
    for part in spec.split(","):
        part = part.strip()
        lo, dash, hi = part.partition("-")
        if dash and lo.strip().isdigit() and hi.strip().isdigit():
            start, end = int(lo), int(hi)
            if start < 1 or start > end:
                raise ValueError(f"page range {part!r} is not ascending from 1 or higher")
            pages.extend(range(start, end + 1))
        elif part.isdigit() and int(part) >= 1:
            pages.append(int(part))
        else:
            raise ValueError(
                f'pages must be "grid", or page numbers like "3", "2-5", "1,4,7" (got {spec!r})'
            )
    return list(dict.fromkeys(pages))


def _refit_slides_to_page(
    data: dict, old_w: float, old_h: float, new_w: float, new_h: float
) -> None:
    """Re-project existing free elements onto a page with another size.

    Filling ``page`` changes the coordinate space the percent geometry
    refers to; leaving the numbers behind silently stretched every shape
    (a circle became ratio 0.750 on a 4:3 skin). This applies the same
    uniform-scale + center letterbox the exporter uses — at save time, so
    the stored deck already means the same picture on the new page.
    Structured slides (title / bullets, no elements) stay untouched: their
    derived layout is defined in page percent and redraws for any ratio.
    Font sizes ride the same uniform scale as the geometry — the exporter
    treats px as an absolute size in the deck's coordinate space, so
    scaling both keeps this exactly the file the old exporter-side
    projection produced for a page-less deck.

    The projection is the direct old-to-new letterbox, and deliberately
    NOT round-trip exact: min-scale letterboxing is not invertible, and
    slide software asks the user how to re-fit on a page change for the
    same reason. What IS guaranteed is containment — coordinates inside
    the old page land inside the new page, always. (An earlier version
    composed through the classic canvas to make swaps path-independent;
    that silently pushed content re-placed for the current page off the
    page entirely. Do not bring that back — off-page content is the worse
    failure.)
    """
    scale = min(new_w / old_w, new_h / old_h)
    offset_x = (new_w - old_w * scale) / 2.0
    offset_y = (new_h - old_h * scale) / 2.0
    font_factor = scale
    for slide in data.get("slides") or []:
        if not isinstance(slide, dict):
            continue
        elements = slide.get("elements")
        if not isinstance(elements, list) or not elements:
            continue
        # The exporter applies `padding` in the NEW space too, so solve the
        # stored percents back out of the projected on-page position.
        pad = (slide.get("padding") or 0.0) / 100.0
        span = 1.0 - 2.0 * pad
        if span <= 0:
            # padding >= 50 leaves no content area to solve back into — the
            # schema refuses it and the layout check names it; never divide
            # by it (a re-fit crash would take the whole save down).
            continue
        for el in elements:
            if not isinstance(el, dict):
                continue
            try:
                ex = pad + (float(el["x"]) / 100.0) * span
                ey = pad + (float(el["y"]) / 100.0) * span
                ew = (float(el["w"]) / 100.0) * span
                eh = (float(el["h"]) / 100.0) * span
            except (KeyError, TypeError, ValueError):
                continue  # malformed element — the exporter reports it later
            el["x"] = round(((offset_x + ex * old_w * scale) / new_w - pad) / span * 100.0, 4)
            el["y"] = round(((offset_y + ey * old_h * scale) / new_h - pad) / span * 100.0, 4)
            el["w"] = round(ew * old_w * scale / new_w / span * 100.0, 4)
            el["h"] = round(eh * old_h * scale / new_h / span * 100.0, 4)
            font = el.get("fontSize", el.get("font_size"))
            if isinstance(font, (int, float)):
                key = "fontSize" if "fontSize" in el else "font_size"
                el[key] = round(font * font_factor, 4)


def _deck_with_skin_page(
    store: CanvasStore, canvas_id: str, content: str
) -> tuple[str, str | None, str | None]:
    """The deck content with ``page`` filled from its template skin.

    A deck that names a template gets the skin's real page size written
    into ``data.page``, so the editor, the preview, and the export agree
    on one aspect ratio — the agent never types the numbers by hand. The
    skin decides the page on a swap too: a deck whose page no longer
    matches its template is re-fitted from its CURRENT page, otherwise the
    content stays stranded in the old ratio's letterbox with no way out
    through coordinates alone. When the size changes the existing elements
    are re-fitted (see :func:`_refit_slides_to_page`); when the RATIO
    changes the third return value carries a note to relay — the change is
    safe but the content now sits letterboxed, and re-placing it is the
    model's call, never a silent one. A deck without a template keeps
    whatever page it has. Content that is not a template-bearing
    envelope passes through untouched (the exporter raises its own honest
    errors later). Returns ``(content, error, note)``: a skin that trips
    the archive safety limits is an error to relay, not a detail to
    absorb.
    """
    try:
        envelope = json.loads(content)
    except json.JSONDecodeError:
        return content, None, None
    data = envelope.get("data") if isinstance(envelope, dict) else None
    if not isinstance(data, dict):
        return content, None, None
    template = data.get("template")
    if not isinstance(template, str) or not template.lower().endswith(".pptx"):
        return content, None, None
    ref = normalize_asset_reference(template)
    if ref is None:
        return content, None, None
    try:
        raw = store.read_bytes(canvas_id, ref).data
    except CanvasStoreError:
        return content, None, None  # missing skin degrades at export time too
    try:
        ensure_archive_within_limits(raw, path=ref)
    except UnsafeArchiveError as exc:
        return content, f"Error: {exc}", None
    size = pptx_page_size_inches(raw)
    if size is None:
        return content, None, None
    new_w, new_h = round(size[0], 4), round(size[1], 4)
    # "The skin decides the page" holds on a swap too, not only on the first
    # attach: the deck's current page is the OLD coordinate space to re-fit
    # from. A deck whose page already matches the skin passes untouched.
    old_w, old_h = DEFAULT_SLIDE_PAGE_IN
    page = data.get("page")
    if isinstance(page, dict):
        got_w = page.get("widthIn", page.get("width_in"))
        got_h = page.get("heightIn", page.get("height_in"))
        if (
            isinstance(got_w, (int, float))
            and isinstance(got_h, (int, float))
            and got_w > 0
            and got_h > 0
        ):
            if abs(got_w - new_w) < 1e-4 and abs(got_h - new_h) < 1e-4:
                return content, None, None
            old_w, old_h = float(got_w), float(got_h)
    data["page"] = {"widthIn": new_w, "heightIn": new_h}
    note = None
    if abs(new_w - old_w) > 1e-4 or abs(new_h - old_h) > 1e-4:
        # Any size change re-fits: with the same ratio the percents come out
        # unchanged and only font sizes ride the physical scale, so the
        # stored deck exports exactly like the un-swapped one did.
        _refit_slides_to_page(data, old_w, old_h, new_w, new_h)
    if abs(new_w / new_h - old_w / old_h) > 1e-6:
        note = (
            f" Note: page changed to {new_w} x {new_h} in to match the "
            "template. Existing slides were scaled to fit (they now sit "
            "letterboxed); re-place their elements to use the full page."
        )
    return json.dumps(envelope, ensure_ascii=False), None, note


def _skin_baseline(
    store: CanvasStore, canvas_id: str, data: dict[str, Any]
) -> tuple[float | None, float, dict[tuple[float, float, float, float], float]]:
    """``(min font px, max overhang, own overflows)`` from the deck's skin.

    A deck copied from an upload is judged by what its author did: the
    smallest size they printed is the readability floor, how far their own
    shapes reach past the page is the overflow allowance, and the boxes
    their own text already overflowed are folded rather than repeated.
    Without a skin the defaults apply.
    """
    template = data.get("template")
    if not isinstance(template, str) or not template.lower().endswith(".pptx"):
        return None, 0.0, {}
    ref = normalize_asset_reference(template)
    if ref is None:
        return None, 0.0, {}
    try:
        raw = store.read_bytes(canvas_id, ref).data
    except CanvasStoreError:
        return None, 0.0, {}
    from .pptx_import import deck_baseline

    baseline = deck_baseline(raw)
    if baseline is None:
        return None, 0.0, {}
    return baseline.smallest_text_px, baseline.max_overhang, baseline.overflow


#: The deck-check lines that mean content is hidden or missing — the ones the
#: export gate refuses on. Small type and layout advice pass; a person can
#: read a 12px footnote, not a line drawn past the box's edge.
_HIDING_FINDINGS = ("run past the box", "off the page", "is not on the")


def create_canvas_tools(
    store: CanvasStore,
    *,
    converters: list[SourceConverter] | None = None,
    title_for: Callable[[str], str] | None = None,
    meta_for: Callable[[str], dict[str, Any] | None] | None = None,
) -> list[Any]:
    """Return the four standard canvas tools bound to ``store``.

    ``converters`` render binary source files (uploads under ``sources/``)
    into model-usable content when the agent reads them; defaults to the
    built-in set (see :mod:`langchain_canvas.converters`) and is fully
    replaceable with your own pipeline. ``title_for`` / ``meta_for``
    customize the live broadcast the same way they customize
    :func:`~langchain_canvas.replay.hydrate_events`: map a file path to a
    display title / renderer hints (for example titling slide files from a
    deck manifest). Defaults: the path as title, no hints.
    """
    active_converters = default_converters() if converters is None else converters

    def _broadcast(
        runtime: ToolRuntime, canvas_id: str, path: str, is_new: bool, commit: Commit
    ) -> None:
        # Same silent no-op contract as the Canvas emitter: no writer, no wire.
        writer = getattr(runtime, "stream_writer", None)
        if writer is None or not path.endswith(ARTIFACT_SUFFIXES):
            return
        content = store.read(canvas_id, path, revision=commit.revision).content
        for event in events_for_commit(
            path,
            content,
            is_new=is_new,
            revision=commit.revision,
            description=commit.description,
            title=title_for(path) if title_for else None,
            meta=meta_for(path) if meta_for else None,
        ):
            writer(event)

    def _has_file(canvas_id: str, path: str) -> bool:
        return any(info.path == path for info in store.list_files(canvas_id))

    def _revision_header(canvas_id: str, path: str, revision: str) -> str:
        """``revision: v4`` plus who last changed the file and when.

        The revision alone says nothing about whether the person has been
        here; the actor and the age do, and they come from the log the store
        already keeps.
        """
        last = last_change_line(store, canvas_id, path)
        return f"revision: {revision}" + (f"\n{last}" if last else "")

    def _edit_document(
        runtime: ToolRuntime,
        canvas_id: str,
        path: str,
        old: str,
        new: str,
        description: str,
        revision: str,
    ) -> str:
        """`edit_canvas` over a Word file — same contract, real paragraphs.

        Word splits one visible sentence across runs, so a replacement that
        matched the file's raw XML would almost never be the one the reader
        asked for. This matches the text as read, across run boundaries, and
        keeps the same one-match-only rule the text path has.
        """
        try:
            got = store.read_bytes(canvas_id, path)
        except CanvasFileNotFoundError as exc:
            return f"Error: {exc}. Use list_canvas_files to see available files."
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        try:
            edited = replace_text(got.data, old, new, path=path)
        except (DocumentOpError, MissingDocumentDependencyError) as exc:
            return f"Error: {exc}"
        return _save_document(
            store,
            runtime,
            canvas_id,
            path,
            edited,
            description,
            revision,
            converters=active_converters,
            verb="Edited",
        )

    def _canvas_paths(canvas_id: str) -> set[str] | None:
        try:
            return {info.path for info in store.list_files(canvas_id)}
        except CanvasStoreError:
            return None

    def _is_checked(path: str) -> bool:
        """Whether a save-time check has anything to say about this path."""
        return path.lower().endswith(".slides.json") or is_document_path(path)

    def _save_note(canvas_id: str, path: str, content: str) -> str:
        """Certain-only warnings for a file just saved ('' when clean).

        Free to compute (fields, coordinates and a file list — no render), so
        it rides every save and the model sees a defect the moment it writes
        one. A deck gets the deck check; a document or page gets the
        reference check, which is the defect that reached readers as a broken
        image. See :mod:`langchain_canvas.layout_lint` and
        :mod:`langchain_canvas.document_lint` for the no-false-positives
        contract both keep.
        """
        if path.lower().endswith(".slides.json"):
            return _deck_note(canvas_id, path, content)
        if is_document_path(path):
            on_canvas = _canvas_paths(canvas_id)
            if on_canvas is None:
                return ""
            return format_document_warnings(
                lint_document_content(
                    content, path=path, ref_exists=on_canvas.__contains__
                )
            )
        return ""

    def _deck_refusal(path: str, content: str) -> str | None:
        """The refusal for a deck save that must not land, or ``None``.

        A deck that is not JSON, or that :func:`blocking_deck_findings`
        objects to, would sit on the canvas broken while this tool reported
        success. Nothing is written; the reply carries the fix.
        """
        try:
            envelope = json.loads(content)
        except json.JSONDecodeError as exc:
            return (
                f"Error: {path} was not saved — it is not valid JSON ({exc.msg} at "
                f"line {exc.lineno}). A deck is "
                '{"type": "slides", "title": "...", "data": {"slides": [...]}}.'
            )
        findings = blocking_deck_findings(envelope, path)
        if not findings:
            return None
        lines = "\n".join(f"  - {finding}" for finding in findings)
        return f"Error: {path} was not saved — fix these and write it again:\n{lines}"

    def _with_default_template(canvas_id: str, content: str) -> tuple[str, str]:
        """A new deck with no ``template`` takes the one PowerPoint upload.

        A person who uploads a deck and asks for a new one in its style
        expects its masters back on export, and a model composing the deck
        forgets the pointer more often than it writes it. With exactly one
        ``.pptx`` under ``sources/`` there is nothing to choose, so the tool
        chooses it and says so; ``"template": null`` written on purpose
        opts out. Returns ``(content, note)``.
        """
        envelope = json.loads(content)
        data = envelope.get("data")
        if not isinstance(data, dict) or "template" in data:
            return content, ""
        on_canvas = _canvas_paths(canvas_id) or set()
        decks = sorted(
            item
            for item in on_canvas
            if item.startswith(_SOURCES_PREFIX) and item.lower().endswith(".pptx")
        )
        if len(decks) != 1:
            return content, ""
        data["template"] = decks[0]
        note = (
            f" Note: template set to {decks[0]} (the only PowerPoint upload), so the "
            'export builds on its masters; write "template": null to build on a blank page.'
        )
        return json.dumps(envelope, ensure_ascii=False), note

    def _table_refusal(
        canvas_id: str, path: str, content: str, *, is_new: bool
    ) -> tuple[str | None, str]:
        """``(refusal, content)`` for a table save: refuse, or normalise.

        Three things put a broken grid in front of the person: a file that is
        not the envelope, keys written outside ``data``, and a rewrite of a
        table the person has formatted (its ``sheet``) that drops that state.
        Everything else is normalised through the schema, so ``columns`` and
        ``rows`` are always present for the renderer.
        """
        shape = (
            '{"type": "table", "title": "...", "data": {"columns": [...], "rows": [...]}}'
        )
        try:
            envelope = json.loads(content)
        except json.JSONDecodeError as exc:
            return (
                f"Error: {path} was not saved — it is not valid JSON ({exc.msg} at line "
                f"{exc.lineno}). A table is {shape}."
            ), content
        if not isinstance(envelope, dict):
            return f"Error: {path} was not saved — a table is {shape}.", content
        misplaced = [k for k in ("columns", "rows", "sheet") if k in envelope]
        if misplaced:
            named = ", ".join(f'"{k}"' for k in misplaced)
            return (
                f"Error: {path} was not saved — {named} must sit inside \"data\", not at "
                f"the top level: {shape}."
            ), content
        data = envelope.get("data")
        if not isinstance(data, dict):
            return f"Error: {path} was not saved — \"data\" is missing: {shape}.", content
        try:
            model = TableData.model_validate(data)
        except Exception as exc:  # noqa: BLE001 - pydantic's message is the fix
            return f"Error: {path} was not saved — {exc}", content
        if not is_new:
            try:
                current = json.loads(store.read(canvas_id, path).content)
            except (CanvasStoreError, ValueError):
                current = None
            has_sheet = isinstance(current, dict) and bool(
                (current.get("data") or {}).get("sheet")
            )
            if has_sheet and not model.sheet:
                return (
                    f"Error: {path} was not saved — it holds the person's grid state "
                    "(formatting, merges, formulas), which this rewrite would erase. "
                    "Change cells with write_table_cells instead."
                ), content
        normalised = model.model_dump(by_alias=True, exclude_none=True)
        title = envelope.get("title")
        try:
            content = encode_artifact(
                {
                    "type": "table",
                    "title": title if isinstance(title, str) else display_title(path),
                    "data": normalised,
                },
                path,
            )
        except ValueError as exc:
            return f"Error: {path} was not saved — {exc}", content
        return None, content

    def _baseline(
        canvas_id: str, data: dict[str, Any]
    ) -> tuple[float | None, float, dict[tuple[float, float, float, float], float]]:
        """``(min font px, max overhang, own overflows)`` from the deck's skin.

        A deck copied from an upload is judged by what its author did: the
        smallest size they printed is the readability floor, and how far their
        own shapes reach past the page is the overflow allowance. Without a
        skin the defaults apply.
        """
        return _skin_baseline(store, canvas_id, data)

    def _deck_note(canvas_id: str, path: str, content: str) -> str:
        """The deck check, over the envelope's ``data`` ('' when unreadable)."""
        try:
            envelope = json.loads(content)
        except json.JSONDecodeError:
            return ""
        data = envelope.get("data") if isinstance(envelope, dict) else None
        if not isinstance(data, dict):
            return ""
        on_canvas = _canvas_paths(canvas_id)
        floor, overhang, own_overflow = _baseline(canvas_id, data)
        warnings = lint_slides_data(
            data,
            ref_exists=None if on_canvas is None else on_canvas.__contains__,
            min_text_px=floor,
            max_overhang=overhang,
            known_overflow=own_overflow,
        )
        return format_layout_warnings(warnings)

    def _deck_eye(
        canvas_id: str,
        path: str,
        content: str,
        save_note: str,
        previous_revision: str | None = None,
    ) -> list[dict]:
        """Page images of the slides the check named, or ``[]``.

        Telling the model to look was measured at 14 asks, 1 look. So when
        a save leaves a finding on a slide and a page renderer is mounted,
        the slide arrives with the finding — rendered from the deck as the
        exporter would print it. Quiet saves, decks without a renderer, and
        any render failure add nothing.

        Among the named slides, the ones this save actually changed come
        first — fourteen saves once arrived with the same two untouched
        slides while the slide being written was never shown. The images are
        resized to glance size on the way out: the model reads a finding,
        not a poster.
        """
        if not path.lower().endswith(".slides.json") or "Deck check" not in save_note:
            return []
        flagged = sorted({int(n) for n in re.findall(r"slide (\d+)", save_note)})
        changed = _changed_slides(canvas_id, path, content, previous_revision)
        touched = [n for n in flagged if n in changed] if changed is not None else flagged
        numbers = touched or flagged
        numbers = numbers[:_EYE_MAX_SLIDES]
        if not numbers:
            return []
        stem = _deck_stem(path)
        renderer = _renderer_for(f"{stem}.pptx", active_converters)
        if renderer is None:
            return []
        try:
            from .exporters import SlidesPptxExporter

            printed = SlidesPptxExporter().export(
                inline_slides_assets(content, store, canvas_id), path=path
            )
            converted = renderer.render_pages(printed.data, path=f"{stem}.pptx", pages=numbers)
        except Exception:  # noqa: BLE001 - the eye is a bonus, never a failure
            return []
        return [
            _glance_size(block)
            for block in converted.blocks
            if block.get("type") == "image"
        ]

    def _changed_slides(
        canvas_id: str, path: str, content: str, previous_revision: str | None
    ) -> set[int] | None:
        """1-based numbers of the slides this save changed, or ``None``.

        ``None`` (no previous revision, or anything unreadable) means "no
        idea", and the caller falls back to every flagged slide.
        """
        if previous_revision is None:
            return None
        try:
            before = store.read(canvas_id, path, revision=previous_revision).content
            old_slides = json.loads(before)["data"]["slides"]
            new_slides = json.loads(content)["data"]["slides"]
        except Exception:  # noqa: BLE001 - a diff is a bonus, never a failure
            return None
        if not isinstance(old_slides, list) or not isinstance(new_slides, list):
            return None
        changed: set[int] = set()
        for index in range(max(len(old_slides), len(new_slides))):
            a = old_slides[index] if index < len(old_slides) else None
            b = new_slides[index] if index < len(new_slides) else None
            if json.dumps(a, sort_keys=True) != json.dumps(b, sort_keys=True):
                changed.add(index + 1)
        return changed

    def _read_source_pages(canvas_id: str, path: str, spec: str) -> str | list[dict]:
        """Rendered page images (or the grid overview) for one paged source.

        Renders are recomputed on every call by design — a persistent
        preview belongs to the file-artifact track, not the read tool.
        """
        converter = converter_for(path, active_converters)
        if converter is None or not isinstance(converter, PageRenderable):
            renderable = sorted(
                {
                    suffix
                    for c in active_converters
                    if isinstance(c, PageRenderable)
                    for suffix in c.suffixes
                }
            )
            supported = ", ".join(renderable) if renderable else "none installed"
            return (
                f"Error: `pages` applies to page-renderable sources ({supported}); "
                f"{path} is not one — read it without `pages` for the text view."
            )
        try:
            got = store.read_bytes(canvas_id, path)
        except CanvasFileNotFoundError as exc:
            return f"Error: {exc}. Use list_canvas_files to see available files."
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        try:
            if spec.strip().lower() == "grid":
                converted = converter.render_grid(got.data, path=path)
            else:
                numbers = _parse_pages(spec)
                if len(numbers) > MAX_IMAGES_PER_CALL:
                    return (
                        f"Error: asked for {len(numbers)} pages; the limit is "
                        f"{MAX_IMAGES_PER_CALL} per call — request a narrower range "
                        '(or pages="grid" for a thumbnail overview).'
                    )
                converted = converter.render_pages(got.data, path=path, pages=numbers)
        except MissingConverterDependencyError as exc:
            return f"Error: {exc}"
        except ValueError as exc:
            return f"Error: {exc}"
        meta = ", ".join(f"{k}: {v}" for k, v in converted.metadata.items())
        header = f"revision: {got.revision}\n{path}" + (f" ({meta})" if meta else "")
        return [{"type": "text", "text": header}, *converted.blocks]

    def _read_source(canvas_id: str, path: str, offset: int, limit: int) -> str | list[dict]:
        """A binary file rendered through its converter (or an honest refusal)."""
        converter = converter_for(path, active_converters)
        if converter is None:
            suffixes = sorted({s for c in active_converters for s in c.suffixes})
            return (
                f"Error: {path} is a binary file and no converter handles it. "
                f"Converters are installed for: {', '.join(suffixes)}."
            )
        got = store.read_bytes(canvas_id, path)
        try:
            converted = converter.convert(got.data, path=path)
        except MissingConverterDependencyError as exc:
            return f"Error: {exc}"
        except UnsafeArchiveError as exc:
            return f"Error: {exc}"
        text = "\n".join(
            str(block.get("text", "")) for block in converted.blocks if block.get("type") == "text"
        )
        sliced, note = _sliced(text, offset, limit)
        meta = ", ".join(f"{k}: {v}" for k, v in converted.metadata.items())
        header = f"{_revision_header(canvas_id, path, got.revision)}\nconverted view of {path}" + (
            f" ({meta})" if meta else ""
        )
        way_in = _way_to_edit(canvas_id, path)
        if way_in:
            header += f"\n{way_in}"
        body = f"{header}\n{sliced}" + (f"\n{note}" if note else "")
        images = [block for block in converted.blocks if block.get("type") == "image"]
        if images:
            return [{"type": "text", "text": body}, *images]
        return body

    def _way_to_edit(canvas_id: str, path: str) -> str:
        """One line under an upload's read: the file it is edited through.

        A read that shows the words but not the door leaves the model to
        guess, and the guess was a fresh file with none of the original's
        formatting. Empty for files that are not uploads.
        """
        if not path.startswith(_SOURCES_PREFIX):
            return ""
        lowered = path.lower()
        if lowered.endswith(".xlsx"):
            copy = working_copy_path(path)
            if _has_file(canvas_id, copy):
                return (
                    f'Editable working copy: {copy} — read it with sheet="s0" and change '
                    "cells with write_table_cells (this upload is read-only)."
                )
            return "This upload is read-only; there is no editable copy on the canvas yet."
        if lowered.endswith(".pptx"):
            return (
                "To edit: open_deck_for_editing makes an editable copy "
                "(this upload is read-only)."
            )
        if lowered.endswith(DOCUMENT_OP_SUFFIXES):
            return (
                "To edit: open_document_for_editing makes an editable copy "
                "(this upload is read-only)."
            )
        return ""

    @tool
    def read_canvas(
        path: str,
        runtime: ToolRuntime,
        offset: int = 0,
        limit: int = _DEFAULT_READ_LIMIT,
        pages: str | None = None,
        sheet: str | None = None,
        fields: str | None = None,
    ) -> str | list[dict]:
        """Read one canvas file before viewing or editing it.

        Returns the file with line numbers plus the current `revision`. You
        need that revision to call `edit_canvas` or to safely overwrite with
        `write_canvas` — always read a file again right before editing it, so
        you see edits the user may have made by hand.

        Long files are windowed: `offset`/`limit` select a line range and the
        output says how to read the rest. Binary uploads under `sources/` are
        rendered through a format converter instead of raw bytes.

        Page-renderable sources ({page_formats}) can also be *seen*: observe cheaply
        first — the default text view names the page count and which pages
        have no text layer — then `pages="grid"` for a one-shot thumbnail
        overview of every page, then `pages="3"` / `"2-5"` / `"1,4,7"` to
        render just the pages that matter (scans, charts, layout questions)
        as images. At most 8 page images per call.

        A `.table.json` table answers with its map, not its contents: every
        rectangle addressed, sized, and counted. A table carries the same
        data twice — `rows`, which is yours, and the grid sheets, which are
        the person's — and the grid is where the size is. Read one with
        `sheet="rows"` or `sheet="s0"`; `offset`/`limit` window that one.

        A `.slides.json` deck can be read as a *projection*: pass
        `fields="color,fontSize,fontFamily"` (any comma-separated element
        keys — type, text, x, y, w, h, fontSize, bold, color, align, fill,
        stroke, strokeWidth, fontFamily, lineHeight, autofit, ... — plus
        slide keys like background, layout) to get one compact line per
        element with just those values. That is how to see a deck's style
        at a glance — check the outline's `colors:`/`fonts:` lines first,
        then project the slide you are about to touch and give a new
        element its neighbours' values instead of inventing new ones.
        """
        canvas_id = _canvas_id(runtime)
        if pages is not None:
            return _read_source_pages(canvas_id, path, pages)
        if sheet is not None and not path.endswith(".table.json"):
            return (
                f"Error: `sheet` applies to .table.json tables; {path} is not "
                "one — read it without `sheet`."
            )
        offset = max(0, offset)
        limit = max(1, limit)
        try:
            got = store.read(canvas_id, path)
        except BinaryContentError:
            try:
                return _read_source(canvas_id, path, offset, limit)
            except CanvasStoreError as exc:
                return f"Error: {exc}."
        except CanvasFileNotFoundError as exc:
            return f"Error: {exc}. Use list_canvas_files to see available files."
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        if path.endswith(".table.json"):
            try:
                view = table_view(got.content, sheet)
            except ValueError as exc:
                return f"Error: {exc}."
            if view is not None and sheet is None:
                return f"{_revision_header(canvas_id, path, got.revision)}\n{view}"
            if view is not None:
                header, _, body = view.partition("\n")
                sliced, note = _sliced(body, offset, limit)
                return (
                    f"{_revision_header(canvas_id, path, got.revision)}\n{header}\n{sliced}"
                    + (f"\n{note}" if note else "")
                )
        if fields is not None:
            if not path.lower().endswith(".slides.json"):
                return (
                    f"Error: `fields` applies to .slides.json decks; {path} is not "
                    "one — read it without `fields`."
                )
            projection = deck_projection(got.content, fields)
            return f"{_revision_header(canvas_id, path, got.revision)}\n{projection}"
        sliced, note = _sliced(got.content, offset, limit)
        header = _revision_header(canvas_id, path, got.revision)
        if path.lower().endswith(".slides.json") and offset == 0:
            outline = deck_outline(got.content)
            if outline:
                header += f"\n{outline}"
        return f"{header}\n{sliced}" + (f"\n{note}" if note else "")

    @tool
    def write_canvas(
        path: str,
        content: str,
        description: str,
        runtime: ToolRuntime,
        revision: str | None = None,
    ) -> str | list[dict]:
        """Create a new canvas file, or fully replace an existing one.

        `description` becomes the version-history entry — one short sentence
        describing the change. For small changes to an existing file prefer
        `edit_canvas`; use `write_canvas` for new files or full rewrites.
        When replacing an existing file, pass the `revision` from your most
        recent `read_canvas` — if the canvas changed since (for example the
        user edited it by hand), the call is rejected instead of silently
        overwriting their work. Omit `revision` only for brand-new files.
        Files under `sources/` (the user's uploads) are read-only.

        Images already on the canvas embed by relative path — an .html page
        uses `<img src="sources/photo.png">` (or `assets/...`), a document
        uses `![photo](sources/photo.png)`. Use the path exactly as
        list_canvas_files shows it, even from a file inside a folder (never
        `../`). The canvas shows them live and exports inline the bytes, so
        never copy an upload to reference it.

        Structured files carry an envelope: a `.slides.json` deck is
        `{"type": "slides", "title": "...", "data": {"slides": [...]}}`.
        Write each slide one of two ways, never both — `elements` is drawn
        instead of the structured fields, not on top of them:

        - Structured, and it is laid out for you: `title`, `subtitle`,
          `bullets` (a list of strings), `bullets2` for the right-hand
          column, `image`, and `layout`. `layout` is exactly one of
          `content` (the default — heading over bullets), `title`,
          `section`, `image`, `two-column`, `blank`. Any other value is
          rejected; omit it rather than invent one. Sizes and positions are
          chosen from the content, so do not add coordinates here.
        - Free `elements`, for a slide you are composing yourself. Every
          element needs an `id` (a short unique string), a `type` of
          text|image|shape|table, and `x`/`y`/`w`/`h` as percent of the
          slide, 0-100. Colors are `#hex`. `fontSize` is px on a 960x540
          page — the layout's own scale is 48 (a cover line), 38 (a
          heading), and 30 / 24 / 19 (body). Picking from it keeps a
          slide you placed by hand next to the ones you did not, and
          keeps a deck from carrying eight sizes nobody chose. Nothing
          under 14px is readable on the canvas or a projector.
          A text box's `autofit` says what happens when the words outgrow
          it: `shape` grows the box to hold them, `text` shrinks the type
          to fit, `none` (the default) leaves the overflow for the deck
          check to name.
          A `table` is `{"id": "t1", "type": "table", "x": 10, "y": 25,
          "w": 80, "h": 40, "rows": [["Item", "Q1"], ["Sales", "120"]],
          "header": true, "stroke": "#9E9E9E", "fontSize": 18}` — the
          words are `rows` (every row the same length), `stroke` draws
          the grid, `colWidths` / `rowHeights` are percent of the table's
          box, and `cells` lists what single cells do differently:
          `{"r": 0, "c": 0, "fill": "#DDEEFF", "bold": true, "colSpan":
          2}`. To change a cell, change its string in `rows`.

        Optional per slide: `background` (a `#hex` string), `notes`.
        Optional `template`, inside `data` next to `slides` and never at
        the top level — `{"type": "slides", "title": "...", "data":
        {"template": "sources/brand.pptx", "slides": [...]}}` — makes the
        pptx export build on that file's masters and layouts. A new deck
        with no `template` takes the only `.pptx` under `sources/` when
        there is exactly one; write `"template": null` to build on a blank
        page instead.

        A deck is refused, not saved, when a deck key sits outside `data`,
        when it does not match the schema, or when a slide carries both
        `elements` and `title`/`bullets` — the reply names what to fix.
        To revise an uploaded deck, do not write a new one: call
        open_deck_for_editing and change the copy with `set_slide_texts`.

        A `.table.json` table is `{"type": "table", "title": "...", "data":
        {"columns": [{"key": "name", "label": "Name"}], "rows": [{"name":
        "Kim"}]}}` — write `columns` and `rows`; `sheet` is the grid editor's
        own state and is never written here. An uploaded spreadsheet already
        has its working copy on the canvas (`<name>.table.json`): change its
        cells with write_table_cells, never rewrite it with this tool — a
        rewrite that drops `sheet` is refused because it would erase the
        person's formatting and formulas.
        """
        if path.startswith(_SOURCES_PREFIX):
            return _sources_readonly(path)
        canvas_id = _canvas_id(runtime)
        is_new = not _has_file(canvas_id, path)
        page_note = None
        template_note = ""
        if path.lower().endswith(".slides.json"):
            refusal = _deck_refusal(path, content)
            if refusal is not None:
                return refusal
            if is_new:
                content, template_note = _with_default_template(canvas_id, content)
            content, page_error, page_note = _deck_with_skin_page(
                store, canvas_id, content
            )
            if page_error is not None:
                return page_error
        elif path.lower().endswith(TABLE_SUFFIX):
            refusal, content = _table_refusal(canvas_id, path, content, is_new=is_new)
            if refusal is not None:
                return refusal
        try:
            commit = store.write(
                canvas_id,
                path,
                content,
                description,
                base_revision=revision,
                actor="agent",
            )
        except RevisionMismatchError as exc:
            return f"Error: {exc}. {_RETRY_HINT}"
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        _broadcast(runtime, canvas_id, path, is_new, commit)
        save_note = _save_note(canvas_id, path, content)
        return _with_eye(
            f"Wrote {path} (revision {commit.revision})."
            f"{template_note}{page_note or ''}{save_note}",
            _deck_eye(canvas_id, path, content, save_note, previous_revision=revision),
        )

    @tool
    def edit_canvas(
        path: str,
        old: str,
        new: str,
        description: str,
        revision: str,
        runtime: ToolRuntime,
    ) -> str | list[dict]:
        """Replace exactly one occurrence of `old` with `new` in a canvas file.

        `revision` must be the value returned by your most recent
        `read_canvas` of this file — if the file changed since (for example
        the user edited it), the call is rejected and you must read again.
        `old` must match exactly once; include enough surrounding context to
        make it unique. `description` is the version-history entry. Files
        under `sources/` (the user's uploads) are read-only.

        Word files ({document_formats}) are edited the same way: `old` is text
        copied from `read_canvas`, matched across the runs Word split it into,
        and may lead with the address the read printed — `"[p7] Title"` — to
        pick that paragraph when the same words appear twice (`"[p7]"` alone
        means the whole paragraph). `new` is the replacement text only; an
        address in front of it is dropped, never written into the document.
        and it must still match exactly once in the whole file — body, tables,
        headers and footers included.
        """
        if path.startswith(_SOURCES_PREFIX):
            return _sources_readonly(path)
        canvas_id = _canvas_id(runtime)
        if old == new:
            return "Error: old and new are the same — nothing to change, nothing saved."
        if _is_document_file(path):
            return _edit_document(runtime, canvas_id, path, old, new, description, revision)
        try:
            commit = store.edit(
                canvas_id,
                path,
                old,
                new,
                description,
                base_revision=revision,
                actor="agent",
            )
        except RevisionMismatchError as exc:
            return f"Error: {exc}. {_RETRY_HINT}"
        except EditConflictError as exc:
            return f"Error: {exc}. {_RETRY_HINT}"
        except CanvasFileNotFoundError as exc:
            return f"Error: {exc}. Use list_canvas_files to see available files."
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        _broadcast(runtime, canvas_id, path, is_new=False, commit=commit)
        save_note = ""
        if _is_checked(path):
            # Small edits introduce defects as easily as full writes — an
            # edit is how a document picks up a reference to a file that is
            # not there. Check the file as the edit left it.
            edited = store.read(canvas_id, path, revision=commit.revision).content
            save_note = _save_note(canvas_id, path, edited)
            return _with_eye(
                f"Edited {path} (revision {commit.revision}).{save_note}",
                _deck_eye(canvas_id, path, edited, save_note, previous_revision=revision),
            )
        return f"Edited {path} (revision {commit.revision}).{save_note}"

    @tool
    def set_slide_texts(
        path: str,
        slide: int,
        texts: dict[str, str],
        description: str,
        revision: str,
        runtime: ToolRuntime,
    ) -> str | list[dict]:
        """Replace the words of one slide's text elements, by element id.

        `slide` is 1-based, as `read_canvas` prints it, and `texts` maps
        element ids to their new words: `{"e0": "Title", "e2": "Body"}`. Put
        every text change for the slide in one call — one call is one save,
        and the reply carries that slide's check (with its image when a
        renderer is mounted). This is the way to retitle and refill a slide;
        it cannot collide with itself the way parallel `edit_canvas` string
        matches do.

        Words only: fonts, colours and boxes stay. Geometry or styling
        changes go through `edit_canvas`; a table's words live in its `rows`
        (`edit_canvas` too). `revision` must be the value from your most
        recent `read_canvas` of this file.
        """
        if not path.lower().endswith(SLIDES_SUFFIX):
            return f"Error: set_slide_texts edits {SLIDES_SUFFIX} decks (got {path})."
        canvas_id = _canvas_id(runtime)
        try:
            got = store.read(canvas_id, path)
        except CanvasFileNotFoundError as exc:
            return f"Error: {exc}. Use list_canvas_files to see available files."
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        try:
            envelope = json.loads(got.content)
            slides = envelope["data"]["slides"]
            assert isinstance(slides, list)
        except Exception:  # noqa: BLE001 - not a deck we can address into
            return f"Error: {path} does not parse as a slides deck; read it and use edit_canvas."
        if not texts:
            return "Error: `texts` is empty — nothing to change."
        if not isinstance(slide, int) or not 1 <= slide <= len(slides):
            return f"Error: slide {slide} is out of range — the deck has {len(slides)} slide(s)."
        elements = slides[slide - 1].get("elements")
        if not isinstance(elements, list) or not elements:
            return (
                f"Error: slide {slide} has no `elements` (it is a structured slide) — "
                "edit its `title`/`bullets` with edit_canvas."
            )
        by_id = {e.get("id"): e for e in elements if isinstance(e, dict)}
        for element_id, words in texts.items():
            element = by_id.get(element_id)
            if element is None:
                have = ", ".join(f'{e.get("id")} ({e.get("type")})' for e in elements)
                return f"Error: slide {slide} has no element {element_id!r}. It has: {have}."
            if element.get("type") == "table":
                return (
                    f"Error: {element_id!r} on slide {slide} is a table — its words are "
                    "its `rows`; change them with edit_canvas."
                )
            if element.get("type") != "text":
                kind = element.get("type")
                return f"Error: {element_id!r} on slide {slide} is a {kind}, not text."
            if not isinstance(words, str):
                return f"Error: the text for {element_id!r} must be a string."
        for element_id, words in texts.items():
            by_id[element_id]["text"] = words
        content = encode_slides(envelope.get("title") or display_title(path), envelope["data"])
        refusal = _deck_refusal(path, content)
        if refusal is not None:
            return refusal
        try:
            commit = store.write(
                canvas_id, path, content, description, base_revision=revision, actor="agent"
            )
        except RevisionMismatchError as exc:
            return f"Error: {exc}. {_RETRY_HINT}"
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        _broadcast(runtime, canvas_id, path, is_new=False, commit=commit)
        save_note = _save_note(canvas_id, path, content)
        return _with_eye(
            f"Set {len(texts)} text(s) on slide {slide} of {path} "
            f"(revision {commit.revision}).{save_note}",
            _deck_eye(canvas_id, path, content, save_note, previous_revision=revision),
        )

    @tool
    def review_deck(
        path: str, runtime: ToolRuntime, slide: int | None = None
    ) -> str | list[dict]:
        """Look over the whole deck before handing it off — check plus pages.

        With no `slide`: the full deck check (the original's own findings
        folded out), a per-slide note of what changed against the original,
        and — when a page renderer is mounted — the deck now as a page grid,
        then the original's grid, for side-by-side comparison. With
        `slide=N`: that one slide rendered large, the deck now first, the
        original second.

        Call it after the last text change and before `export_canvas` — the
        export refuses while findings that hide content remain.
        """
        canvas_id = _canvas_id(runtime)
        if not path.lower().endswith(SLIDES_SUFFIX):
            return f"Error: review_deck reads {SLIDES_SUFFIX} decks (got {path})."
        try:
            content = store.read(canvas_id, path).content
        except CanvasFileNotFoundError as exc:
            return f"Error: {exc}. Use list_canvas_files to see available files."
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        try:
            data = json.loads(content)["data"]
            slides = data["slides"]
            assert isinstance(slides, list)
        except Exception:  # noqa: BLE001 - not a reviewable deck
            return f"Error: {path} does not parse as a slides deck."
        if slide is not None and not 1 <= slide <= len(slides):
            return f"Error: slide {slide} is out of range — the deck has {len(slides)} slide(s)."

        note = _save_note(canvas_id, path, content) or "\nDeck check: clean."
        lines = [f"{path} — {len(slides)} slide(s).{note}"]

        # What moved against the original, slide by slide — the review is
        # where "slide 5 was never touched" stops being invisible.
        original = _original_deck_data(canvas_id, data)
        if original is not None:
            source_slides = original.get("slides") or []
            for number, current in enumerate(slides, start=1):
                if slide is not None and number != slide:
                    continue
                base = source_slides[number - 1] if number - 1 < len(source_slides) else {}
                base_texts = {
                    e.get("id"): e.get("text")
                    for e in (base.get("elements") or [])
                    if isinstance(e, dict) and e.get("type") == "text"
                }
                changed, untouched = [], []
                for element in current.get("elements") or []:
                    if not isinstance(element, dict) or element.get("type") != "text":
                        continue
                    before = base_texts.get(element.get("id"))
                    if before is None:
                        changed.append(element.get("id"))
                    elif element.get("text") == before:
                        untouched.append(element.get("id"))
                    else:
                        changed.append(element.get("id"))
                parts = [f"changed {len(changed)} text(s)"]
                if untouched:
                    shown = ", ".join(str(i) for i in untouched[:4])
                    parts.append(
                        f"{len(untouched)} still read as the original ({shown}"
                        + (", ..." if len(untouched) > 4 else "")
                        + ")"
                    )
                lines.append(f"[s{number}] " + "; ".join(parts))

        images = _review_images(canvas_id, path, content, data, slide)
        if images:
            what = "that slide" if slide is not None else "every page"
            lines.append(f"Images: {what} of the deck now, then the original.")
        return _with_eye("\n".join(lines), images)

    def _original_deck_data(canvas_id: str, data: dict[str, Any]) -> dict[str, Any] | None:
        """The template re-imported, for comparison — or ``None``, quietly."""
        template = data.get("template")
        if not isinstance(template, str) or not template.lower().endswith(".pptx"):
            return None
        ref = normalize_asset_reference(template)
        if ref is None:
            return None
        try:
            raw = store.read_bytes(canvas_id, ref).data
            from .pptx_import import pptx_to_slides

            return pptx_to_slides(raw)
        except Exception:  # noqa: BLE001 - the comparison is a bonus
            return None

    def _review_images(
        canvas_id: str,
        path: str,
        content: str,
        data: dict[str, Any],
        slide: int | None,
    ) -> list[dict]:
        """Current pages then original pages, glance-sized — or ``[]``."""
        stem = _deck_stem(path)
        renderer = _renderer_for(f"{stem}.pptx", active_converters)
        if renderer is None:
            return []
        try:
            from .exporters import SlidesPptxExporter

            printed = SlidesPptxExporter().export(
                inline_slides_assets(content, store, canvas_id), path=path
            )
            blocks: list[dict] = []
            template = data.get("template")
            ref = normalize_asset_reference(template) if isinstance(template, str) else None
            original = store.read_bytes(canvas_id, ref).data if ref else None
            if slide is not None:
                converted = renderer.render_pages(
                    printed.data, path=f"{stem}.pptx", pages=[slide]
                )
                blocks += converted.blocks
                if original is not None:
                    converted = renderer.render_pages(
                        original, path=f"{stem}.pptx", pages=[slide]
                    )
                    blocks += converted.blocks
            else:
                blocks += renderer.render_grid(printed.data, path=f"{stem}.pptx").blocks
                if original is not None:
                    blocks += renderer.render_grid(original, path=f"{stem}.pptx").blocks
            return [
                _glance_size(block) for block in blocks if block.get("type") == "image"
            ]
        except Exception:  # noqa: BLE001 - the review still reports in text
            return []

    @tool
    def list_canvas_files(runtime: ToolRuntime) -> str:
        """List the files currently on the canvas, with sizes in bytes."""
        try:
            infos = store.list_files(_canvas_id(runtime))
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        if not infos:
            return "The canvas is empty."
        return "\n".join(f"{info.path} ({info.size} bytes)" for info in infos)

    # The eye reaches whatever page-renderable converters are installed, and
    # a format the description does not name is a format the model will say
    # it cannot see. An agent asked to look at page 2 of an imported deck
    # answered "I have no tool for that" while the tool sat in its own list.
    renderable = sorted(
        {
            suffix
            for converter in active_converters
            if isinstance(converter, PageRenderable)
            for suffix in converter.suffixes
        }
    )
    read_canvas.description = read_canvas.description.replace(
        "{page_formats}", ", ".join(renderable) if renderable else "none installed"
    )
    edit_canvas.description = edit_canvas.description.replace(
        "{document_formats}", _DOCUMENT_FORMATS
    )
    return [read_canvas, write_canvas, edit_canvas, set_slide_texts, review_deck, list_canvas_files]


def create_document_tools(
    store: CanvasStore, *, converters: list[SourceConverter] | None = None
) -> list[Any]:
    """Build the Word-editing tools bound to ``store``.

    The agent already replaces text in a document through ``edit_canvas``;
    these are the edits text replacement cannot express — adding a paragraph,
    dropping one, swapping a picture — plus the copy step that makes an
    upload editable in the first place. Kept out of
    :func:`create_canvas_tools` so the four standard tools stay a stable
    contract; mount these when your agent should hand documents back.

    Uploads under ``sources/`` stay read-only here as everywhere else. The
    working copy is what gets edited, and the original the user sent is still
    on the canvas next to it, unchanged, for as long as the canvas lives.

    ``converters`` is the same list :func:`create_canvas_tools` takes, used
    here to prove an edited file still renders before it is saved; defaults
    to the built-in set.
    """
    active_converters = default_converters() if converters is None else converters

    def _document(canvas_id: str, path: str) -> tuple[bytes | None, str]:
        """The stored bytes of a document path, or why they are not usable."""
        if not _is_document_file(path):
            return None, (
                f"Error: these operations edit {_DOCUMENT_FORMATS} files (got {path}). "
                + _other_ways_to_open(path, active_converters)
            )
        if path.startswith(_SOURCES_PREFIX):
            return None, _sources_readonly(path)
        try:
            return store.read_bytes(canvas_id, path).data, ""
        except CanvasFileNotFoundError as exc:
            return None, f"Error: {exc}. Use list_canvas_files to see available files."
        except CanvasStoreError as exc:
            return None, f"Error: {exc}."

    @tool
    def open_document_for_editing(
        source: str, runtime: ToolRuntime, destination: str | None = None
    ) -> str:
        """Make an editable copy of an uploaded Word file.

        Use this once, before the first edit: uploads under `sources/` are the
        user's originals and stay read-only, so edits go to a copy. The copy
        lands at the canvas root named `Editing - <file name>`, so the two sit
        side by side and the user can tell at a glance which one is theirs and
        which one you are changing. The marker goes in front because the two
        names are otherwise identical and a long one is clipped at the end.
        Pass `destination` to name the copy yourself.

        After this, read the copy with `read_canvas` and change it with
        `edit_canvas`, `insert_document_paragraph`, `insert_document_image`,
        `remove_document_paragraph` and `replace_document_image`.
        """
        canvas_id = _canvas_id(runtime)
        if not _is_document_file(source):
            return (
                f"Error: this opens {_DOCUMENT_FORMATS} files (got {source}). "
                + _other_ways_to_open(source, active_converters)
            )
        target = (destination or _working_copy_name(source)).strip()
        if target.startswith(_SOURCES_PREFIX):
            return (
                "Error: sources/ holds the user's uploads — put the working copy "
                "somewhere else (the default is the marked name at the canvas root)."
            )
        if not _is_document_file(target):
            return f"Error: the copy has to keep the document's format ({_DOCUMENT_FORMATS})."
        try:
            got = store.read_bytes(canvas_id, source)
        except CanvasFileNotFoundError as exc:
            return f"Error: {exc}. Use list_canvas_files to see available files."
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        if any(info.path == target for info in store.list_files(canvas_id)):
            return (
                f"Error: {target} is already on the canvas. Edit that one, or pass "
                "`destination` to copy under another name."
            )
        try:
            commit = store.write_bytes(
                canvas_id, target, got.data, f"Copy {source} for editing", actor="agent"
            )
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        _broadcast_source(store, runtime, canvas_id, target, True, commit)
        return (
            f"Copied {source} to {target} ({len(got.data)} bytes, revision "
            f"{commit.revision}). Edit {target}; {source} keeps the user's original."
        )

    @tool
    def insert_document_paragraph(
        path: str,
        anchor: str,
        text: str,
        description: str,
        revision: str,
        runtime: ToolRuntime,
        style: str | None = None,
        position: str = "after",
    ) -> str:
        """Add a paragraph next to one that is already in a Word file.

        `anchor` is text copied from `read_canvas` that appears exactly once
        in the document — it names the paragraph to add next to, and an anchor
        matching nothing or several places is refused rather than guessed at.
        `position` is `after` (the default) or `before`. `style` names a
        paragraph style the document already has, such as `Heading 2` or
        `List Bullet`; `read_canvas` shows each paragraph's style in
        parentheses, so match the paragraphs around the new one. Omitted, the
        new paragraph takes the document's default style.

        This writes text. A picture goes in with `insert_document_image` —
        markdown image syntax is refused here, because a Word file would show
        it as the characters themselves.

        `revision` must come from your most recent `read_canvas` of this file.
        """
        canvas_id = _canvas_id(runtime)
        found = _MARKDOWN_IMAGE.search(text)
        if found:
            return (
                f"Error: {found.group(0)!r} is markdown image syntax, which a Word file "
                "shows as plain text rather than a picture. Use insert_document_image to "
                "put the picture in."
            )
        data, problem = _document(canvas_id, path)
        if data is None:
            return problem
        try:
            edited = insert_paragraph(
                data, anchor=anchor, text=text, style=style, position=position, path=path
            )
        except (DocumentOpError, MissingDocumentDependencyError) as exc:
            return f"Error: {exc}"
        return _save_document(
            store,
            runtime,
            canvas_id,
            path,
            edited,
            description,
            revision,
            converters=active_converters,
            verb="Added a paragraph to",
        )

    @tool
    def remove_document_paragraph(
        path: str, anchor: str, description: str, revision: str, runtime: ToolRuntime
    ) -> str:
        """Delete one paragraph from a Word file.

        `anchor` is text copied from `read_canvas` that appears exactly once
        in the document; the paragraph holding it is the one removed. A
        paragraph that is the only one in its table cell cannot be removed —
        Word needs one there — so replace its text instead.

        `revision` must come from your most recent `read_canvas` of this file.
        """
        canvas_id = _canvas_id(runtime)
        data, problem = _document(canvas_id, path)
        if data is None:
            return problem
        try:
            edited = remove_paragraph(data, anchor=anchor, path=path)
        except (DocumentOpError, MissingDocumentDependencyError) as exc:
            return f"Error: {exc}"
        return _save_document(
            store,
            runtime,
            canvas_id,
            path,
            edited,
            description,
            revision,
            converters=active_converters,
            verb="Removed a paragraph from",
        )

    @tool
    def insert_document_image(
        path: str,
        image_path: str,
        description: str,
        revision: str,
        runtime: ToolRuntime,
        anchor: str | None = None,
        position: str = "after",
        width_inches: float | None = None,
        alt_text: str | None = None,
    ) -> str:
        """Put a picture into a Word file.

        `image_path` is a file already on the canvas, under `assets/` or
        `sources/`; bring an image the user attached onto the canvas before
        pointing at it here. Nothing is fetched from outside the canvas.

        Leave `anchor` out to put the picture at the end of the document.
        Given one — text copied from `read_canvas` that appears exactly once —
        the picture goes next to that paragraph, `after` it by default or
        `before` it.

        The picture arrives at its own size, brought down to the width of the
        text column if it is wider than that; it is never enlarged. Pass
        `width_inches` to choose the width yourself, and the height follows
        the picture's proportions. `alt_text` describes it for a reader who
        cannot see it.

        `revision` must come from your most recent `read_canvas` of this file.
        """
        canvas_id = _canvas_id(runtime)
        data, problem = _document(canvas_id, path)
        if data is None:
            return problem
        reference = normalize_asset_reference(image_path)
        if reference is None:
            return (
                f"Error: {image_path} is not a canvas image path — pass one under "
                f"{ASSETS_PREFIX} or {_SOURCES_PREFIX}, exactly as list_canvas_files "
                "shows it."
            )
        try:
            picture = store.read_bytes(canvas_id, reference).data
        except CanvasFileNotFoundError as exc:
            return f"Error: {exc}. Use list_canvas_files to see available files."
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        try:
            edited, note = insert_image(
                data,
                image=picture,
                anchor=anchor,
                position=position,
                width_inches=width_inches,
                alt_text=alt_text,
                path=path,
            )
        except (DocumentOpError, MissingDocumentDependencyError) as exc:
            return f"Error: {exc}"
        return _save_document(
            store,
            runtime,
            canvas_id,
            path,
            edited,
            description,
            revision,
            converters=active_converters,
            verb="Added a picture to",
            note=f" {note}",
        )

    @tool
    def replace_document_image(
        path: str,
        index: int,
        image_path: str,
        description: str,
        revision: str,
        runtime: ToolRuntime,
    ) -> str:
        """Swap the picture at `[img<index>]` in a Word file for another image.

        `index` is the number `read_canvas` prints next to the picture —
        `[img0]` is 0. `image_path` is a file already on the canvas, under
        `assets/` or `sources/`; bring an image the user attached onto the
        canvas before pointing at it here. The picture keeps the width it has
        on the page and its height is refitted to the new image, so the
        layout around it does not move.

        `revision` must come from your most recent `read_canvas` of this file.
        """
        canvas_id = _canvas_id(runtime)
        data, problem = _document(canvas_id, path)
        if data is None:
            return problem
        reference = normalize_asset_reference(image_path)
        if reference is None:
            return (
                f"Error: {image_path} is not a canvas image path — pass one under "
                f"{ASSETS_PREFIX} or {_SOURCES_PREFIX}, exactly as list_canvas_files "
                "shows it."
            )
        try:
            picture = store.read_bytes(canvas_id, reference).data
        except CanvasFileNotFoundError as exc:
            return f"Error: {exc}. Use list_canvas_files to see available files."
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        try:
            edited, note = replace_image(data, index=index, image=picture, path=path)
        except (DocumentOpError, MissingDocumentDependencyError) as exc:
            return f"Error: {exc}"
        return _save_document(
            store,
            runtime,
            canvas_id,
            path,
            edited,
            description,
            revision,
            converters=active_converters,
            verb="Replaced a picture in",
            note=f" {note}",
        )

    return [
        open_document_for_editing,
        insert_document_paragraph,
        insert_document_image,
        remove_document_paragraph,
        replace_document_image,
    ]


def _run_formula_evaluator(
    evaluator: Sequence[str], payload: dict[str, Any]
) -> list[dict[str, Any]] | None:
    """The evaluator's ``results`` list, or ``None`` when it could not run."""
    try:
        proc = subprocess.run(  # noqa: S603 — host-configured argv, no shell
            list(evaluator),
            input=json.dumps(payload).encode(),
            capture_output=True,
            timeout=60,
        )
        output = json.loads(proc.stdout.decode())
        results = output["results"]
        return results if isinstance(results, list) else None
    except (OSError, ValueError, KeyError, subprocess.TimeoutExpired):
        return None


def _stamp_sheet_formulas(
    content: str, sheet: str, written: dict[str, Any], evaluator: Sequence[str] | None
) -> tuple[str, str]:
    """Compute the touched sheet's formulas and stamp their values.

    Measured before this existed: an agent-written ``=ROUND(H2*I2*2,0)``
    stored only ``f``, the grid showed a blank cell, and a value someone
    else's formula depended on left the dependent showing its stale cache.
    So after a write the whole sheet is recomputed — dependents included —
    and every formula cell gets the value the person will see (``v``/``m``),
    the same way the importer carries the file's own cached values across.

    Returns ``(content, note)`` — content unchanged and a note that says so
    when there is nothing to compute or no way to compute it.
    """
    from .table_outline import _a1, _at, _grid_index, _rewritten, _table, cell_map

    try:
        envelope, data, sheets = _table(content)
        index = _grid_index(sheet, sheets)
    except ValueError:
        return content, ""
    grid = cell_map(sheets[index])
    formula_cells = {
        at: cell
        for at, cell in grid.items()
        if isinstance(cell, dict) and isinstance(cell.get("f"), str) and cell["f"].strip()
    }
    if not formula_cells:
        return content, ""
    if evaluator is None:
        return content, (
            f"\nFormulas: {len(formula_cells)} formula cell(s) were NOT recomputed "
            "(no evaluator configured) — the grid may show stale or blank values."
        )
    celldata = [{"r": r, "c": c, "v": v} for (r, c), v in sorted(grid.items())]
    results = _run_formula_evaluator(evaluator, {"sheets": [{"celldata": celldata}]})
    if not results:
        return content, (
            f"\nFormulas: {len(formula_cells)} formula cell(s) were NOT recomputed "
            "(the evaluator could not run or predates grid mode) — run check_table."
        )
    written_at = set()
    for ref in written:
        try:
            written_at.add(_at(ref))
        except ValueError:
            continue
    shown: list[str] = []
    errors: list[str] = []
    stamped_others = 0
    for result in results:
        at = (int(result["r"]), int(result["c"]))
        cell = grid.get(at)
        if not isinstance(cell, dict):
            continue
        value = result.get("value")
        address = _a1(*at)
        if value in (None, "#ERR"):
            # Not stamped: a wrong number is worse than a blank one. The old
            # cached value (if any) stays, honestly flagged.
            errors.append(
                f"{address} {cell['f']} → #ERR — " + _error_hint(str(cell["f"]))
            )
            continue
        cell["v"] = value
        cell["m"] = str(value)
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            cell["ct"] = {"fa": "General", "t": "n"}
        elif "ct" in cell:
            cell.pop("ct")
        if at in written_at:
            shown.append(f"{address} {cell['f']} → {value}")
        else:
            stamped_others += 1
    sheets[index]["celldata"] = [
        {"r": r, "c": c, "v": v} for (r, c), v in sorted(grid.items())
    ]
    parts = shown[:8] + errors[:8]
    if stamped_others:
        parts.append(f"{stamped_others} other formula cell(s) recomputed with them")
    note = "\nFormulas: " + " · ".join(parts) if parts else ""
    return _rewritten(envelope), note


def _recalc_with_workbook_engine(
    content: str, xlsx_recalc: Callable[[bytes], bytes]
) -> tuple[str, str]:
    """Stamp every grid formula with a full spreadsheet engine's values.

    Called when the light engine left ``#ERR`` cells: the table exports to
    xlsx, the host's ``xlsx_recalc`` (LibreOffice behind an endpoint, in the
    reference deployment) recalculates the whole workbook, and the computed
    values come back onto the formula cells — the engine that will open the
    exported file is the engine that filled the screen. Any failure returns
    the content unchanged with an honest note.
    """
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]

        envelope = json.loads(content)
        data = envelope.get("data") or {}
        sheets = data.get("sheet") or []
        if not sheets:
            return content, ""
        exported = TableXlsxExporter().export(content, path="recalc.table.json")
        fresh = load_workbook(io.BytesIO(xlsx_recalc(exported.data)), data_only=True)
        stamped: list[str] = []
        from .table_outline import _a1

        for index, sheet in enumerate(sheets):
            if index >= len(fresh.worksheets):
                break
            ws = fresh.worksheets[index]
            for cell in sheet.get("celldata") or []:
                v = cell.get("v")
                if not (isinstance(v, dict) and isinstance(v.get("f"), str) and v["f"].strip()):
                    continue
                value = ws.cell(row=int(cell["r"]) + 1, column=int(cell["c"]) + 1).value
                if value is None:
                    continue
                if not isinstance(value, (int, float, str, bool)):
                    value = str(value)
                had_value = "v" in v
                v["v"] = value
                v["m"] = str(value)
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    v["ct"] = {"fa": "General", "t": "n"}
                elif "ct" in v:
                    v.pop("ct")
                if not had_value:
                    stamped.append(f"{_a1(int(cell['r']), int(cell['c']))} → {value}")
        note = (
            "\nRecalculated with the full spreadsheet engine: "
            + " · ".join(stamped[:8])
            if stamped
            else "\nRecalculated with the full spreadsheet engine."
        )
        return json.dumps(envelope, ensure_ascii=False), note
    except Exception as exc:  # noqa: BLE001 - the save must land either way
        return content, (
            f"\nFull recalculation failed ({exc}) — the flagged cells keep no value "
            "until check_table or the export recalculates them."
        )


def create_table_tools(
    store: CanvasStore,
    *,
    evaluator: Sequence[str] | None = None,
    xlsx_recalc: Callable[[bytes], bytes] | None = None,
) -> list[Any]:
    """Build the table-writing tools bound to ``store``.

    ``read_canvas`` shows a table as a map and hands back one addressed
    rectangle; these write to the same addresses. Without them the only way
    to change a cell is to rewrite the whole file, which on a real import
    means resending millions of characters and losing the person's
    formatting, merges and other sheets in the process.

    Kept out of :func:`create_canvas_tools` so the four standard tools stay a
    stable contract; mount these when your agent should edit spreadsheets.
    """

    def _load(canvas_id: str, path: str) -> Any:
        if not path.endswith(".table.json"):
            if path.startswith(_SOURCES_PREFIX) and path.lower().endswith(".xlsx"):
                raise ValueError(
                    f"{path} is the uploaded original (read-only); its working copy is "
                    f"{working_copy_path(path)} — read that with sheet=\"s0\" and write "
                    "cells there"
                )
            raise ValueError(f"this works on .table.json tables (got {path})")
        return store.read(canvas_id, path)

    def _save(
        runtime: ToolRuntime,
        canvas_id: str,
        path: str,
        content: str,
        description: str,
        revision: str,
        note: str,
    ) -> str:
        try:
            commit = store.write(
                canvas_id, path, content, description, base_revision=revision, actor="agent"
            )
        except RevisionMismatchError as exc:
            return f"Error: {exc}. {_RETRY_HINT}"
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        # Same silent no-op contract as the other writers: no writer, no wire.
        writer = getattr(runtime, "stream_writer", None)
        if writer is not None:
            for event in events_for_commit(
                path,
                content,
                is_new=False,
                revision=commit.revision,
                description=commit.description,
            ):
                writer(event)
        return f"{note} ({path}, revision {commit.revision})"

    @tool
    def write_table_cells(
        path: str,
        sheet: str,
        cells: dict[str, Any],
        description: str,
        revision: str,
        runtime: ToolRuntime,
    ) -> str:
        """Put values into one sheet of a .table.json table, cell by cell.

        `sheet` is an address `read_canvas` printed — `s0`, `s1`, ... — and
        `cells` maps cell addresses to what goes in them:
        `{"B3": 42, "C3": "=B3*2", "D3": "done"}`. The column letters are in
        the `### sheet:` line of the read, so nothing needs counting. A value
        starting with `=` stays a formula; `""` clears the cell. Styling on a
        cell you overwrite stays. On save the sheet's formulas are computed —
        dependents included — and the reply shows each written formula's
        value; a formula the grid cannot run is flagged `#ERR` instead of
        showing a wrong or stale number.

        To match the sheet's look, write a dict: `{"v": "Notes", "like":
        "A3"}` copies A3's bold, fill, font and colour onto the cell; explicit
        keys win over the copy — `{"v": "Total", "bl": 1, "bg": "#DDEBF7",
        "fc": "#1F4E78", "fs": 11}`. The read's `styles:` lines say where each
        look lives, so "like the header" is `like` the header's address.

        This is how to change a table: rewriting the whole file with
        `write_canvas` replaces every sheet and drops the formatting, merges
        and formulas only the grid holds. `revision` must be the value from
        your most recent `read_canvas` of this file.
        """
        canvas_id = _canvas_id(runtime)
        try:
            got = _load(canvas_id, path)
            content, note = table_write_cells(got.content, sheet, cells)
            stamped, formula_note = _stamp_sheet_formulas(content, sheet, cells, evaluator)
            content, note = stamped, note + formula_note
            if xlsx_recalc is not None and "#ERR" in formula_note:
                content, recalc_note = _recalc_with_workbook_engine(content, xlsx_recalc)
                note += recalc_note
        except CanvasFileNotFoundError as exc:
            return f"Error: {exc}. Use list_canvas_files to see available files."
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        except ValueError as exc:
            return f"Error: {exc}."
        return _save(runtime, canvas_id, path, content, description, revision, note)

    @tool
    def add_table_sheet(
        path: str,
        name: str,
        description: str,
        revision: str,
        runtime: ToolRuntime,
    ) -> str:
        """Add an empty sheet to a .table.json table.

        The new sheet goes last and takes the next address, so the addresses
        `read_canvas` already gave you keep pointing at the same sheets.
        Write into it with `write_table_cells`. `revision` must be the value
        from your most recent `read_canvas` of this file.
        """
        canvas_id = _canvas_id(runtime)
        try:
            got = _load(canvas_id, path)
            content, note = table_add_sheet(got.content, name)
        except CanvasFileNotFoundError as exc:
            return f"Error: {exc}. Use list_canvas_files to see available files."
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        except ValueError as exc:
            return f"Error: {exc}."
        return _save(runtime, canvas_id, path, content, description, revision, note)

    return [write_table_cells, add_table_sheet]


def create_export_tool(
    store: CanvasStore,
    *,
    exporters: list[Exporter] | None = None,
    converters: list[SourceConverter] | None = None,
    xlsx_recalc: Callable[[bytes], bytes] | None = None,
) -> Any:
    """Build an ``export_canvas`` tool bound to ``store``.

    ``converters`` is the same list ``create_canvas_tools`` takes; when one of
    them renders pages for the exported format, the reply carries a thumbnail
    grid of the file that was just written — the export is the moment the
    result is final, so it is the moment to look.

    Kept separate from :func:`create_canvas_tools` so the four standard tools
    stay a stable contract — add this tool when your agent should hand users
    office files. ``exporters`` defaults to the built-in set (see
    :mod:`langchain_canvas.exporters`) and is fully replaceable with your own
    pipeline. Exported files land on the canvas under ``exports/``.
    """
    active_exporters = default_exporters() if exporters is None else exporters

    active_converters = default_converters() if converters is None else converters

    def _deck_export_gate(canvas_id: str, content: str) -> str | None:
        """The refusal for a deck that would export with hidden content, or None.

        Only the findings that hide words block (see ``_HIDING_FINDINGS``);
        the template's own inherited overflows are already folded out, so the
        list is one the agent can actually bring to zero.
        """
        try:
            data = json.loads(content).get("data")
        except (ValueError, AttributeError):
            return None  # unparseable decks are the exporter's own refusal
        if not isinstance(data, dict):
            return None
        floor, overhang, overflow = _skin_baseline(store, canvas_id, data)
        try:
            on_canvas = {info.path for info in store.list_files(canvas_id)}
        except CanvasStoreError:
            on_canvas = set()
        findings = lint_slides_data(
            data,
            ref_exists=on_canvas.__contains__,
            min_text_px=floor,
            max_overhang=overhang,
            known_overflow=overflow,
        )
        blocking = [w for w in findings if any(k in w for k in _HIDING_FINDINGS)]
        if not blocking:
            return None
        listed = "\n".join(f"- {w}" for w in blocking[:3])
        more = f"\n- ... and {len(blocking) - 3} more" if len(blocking) > 3 else ""
        return (
            f"Error: not exported — {len(blocking)} finding(s) would hide content "
            "in the file:\n"
            f"{listed}{more}\n"
            "Fix them (set_slide_texts for words, edit_canvas for boxes), see the "
            "pages with review_deck, or — only once the user has said to export "
            "as is — call again with accept_findings=True."
        )

    @tool
    def export_canvas(
        path: str, target: str, runtime: ToolRuntime, accept_findings: bool = False
    ) -> str | list[dict]:
        """Export canvas work into a downloadable office file.

        ``path`` is one canvas file (``report/02-overview.html``,
        ``sales.table.json``) or a directory prefix ending in ``/``
        (``report/``), which merges every .html file under it, in name
        order, into one document with a page break between sections.
        ``target`` is the output format: ``docx`` for .html and .md files,
        ``xlsx`` for .table.json tables, ``pptx`` for .slides.json decks.
        The result is saved under ``exports/`` on the canvas, where the
        user can download it.

        Template skin: a slides deck whose data carries
        ``"template": "sources/brand.pptx"`` exports onto that file's
        masters and layouts, so the original's logos, backgrounds, and
        headers survive, and the text takes the face that file uses most.
        A missing or unreadable skin degrades to the plain blank-layout
        export, where fonts fall back to whatever the viewer has installed.

        A deck is refused while its check still names content-hiding
        findings (text past its box, boxes off the page, missing images) —
        exporting those hands the user a file with invisible words. Fix
        them, look with `review_deck`, or pass `accept_findings=True` only
        after the user has said to export as is.
        """
        canvas_id = _canvas_id(runtime)
        try:
            if path.endswith("/"):
                section_paths = sorted(
                    info.path
                    for info in store.list_files(canvas_id)
                    if info.path.startswith(path) and info.path.lower().endswith((".html", ".htm"))
                )
                if not section_paths:
                    return f"Error: no .html files under {path} to export."
                content = "\n<hr/>\n".join(
                    store.read(canvas_id, section).content for section in section_paths
                )
                sample = section_paths[0]
            else:
                content = store.read(canvas_id, path).content
                sample = path
        except CanvasFileNotFoundError as exc:
            return f"Error: {exc}. Use list_canvas_files to see available files."
        except BinaryContentError:
            return (
                f"Error: {path} is binary; export reads text canvas files "
                "(.html, .table.json, .slides.json)."
            )

        if sample.lower().endswith(".slides.json") and not accept_findings:
            refusal = _deck_export_gate(canvas_id, content)
            if refusal is not None:
                return refusal
        # Relative asset references (assets/, sources/) become data: URIs here,
        # before the exporter runs — exporters keep their one-method contract
        # and the exported file leaves self-contained, images included.
        if sample.lower().endswith((".html", ".htm")):
            content = inline_canvas_assets(content, store, canvas_id)
        elif sample.lower().endswith(".slides.json"):
            content = inline_slides_assets(content, store, canvas_id)

        exporter = exporter_for(sample, target, active_exporters)
        if exporter is None:
            available = sorted(
                {e.target for e in active_exporters if sample.lower().endswith(e.suffixes)}
            )
            hint = f" Formats available for this file: {', '.join(available)}." if available else ""
            return f"Error: no exporter turns {sample} into {target!r}.{hint}"
        try:
            exported = exporter.export(content, path=path)
        except MissingExporterDependencyError as exc:
            return f"Error: {exc}"
        except ValueError as exc:
            return f"Error: {exc}"

        recalc_note = ""
        if xlsx_recalc is not None and exported.filename.lower().endswith(".xlsx"):
            # The file leaves with every cached value freshly computed, so it
            # opens showing numbers — including formulas no lighter engine
            # covers, and whatever a person typed straight into the grid.
            try:
                exported = ExportedFile(
                    xlsx_recalc(exported.data), exported.filename, exported.media_type
                )
            except Exception as exc:  # noqa: BLE001 - the export still lands
                recalc_note = (
                    f" (Formula caches were not recalculated: {exc}; spreadsheet "
                    "apps recalculate on open.)"
                )
        out_path = f"exports/{exported.filename}"
        commit = store.write_bytes(
            canvas_id,
            out_path,
            exported.data,
            f"Export {path} -> {exported.filename}",
            actor="agent",
        )
        reply = (
            f"Exported {path} to {out_path} ({len(exported.data)} bytes, revision "
            f"{commit.revision}). The user can download it from the canvas file "
            f"list.{recalc_note}"
        )
        renderer = _renderer_for(out_path, active_converters)
        if renderer is None:
            return reply
        try:
            converted = renderer.render_grid(exported.data, path=out_path)
        except Exception:  # noqa: BLE001 - the look is a bonus; the export already landed
            return reply + " (The page renderer could not draw it — read it with pages=\"grid\".)"
        images = [block for block in converted.blocks if block.get("type") == "image"]
        if not images:
            return reply
        return _with_eye(
            reply + " These are its pages as exported — check them before telling the user "
            "it is done.",
            images,
        )

    return export_canvas


def create_asset_tool(store: CanvasStore) -> Any:
    """Build a ``write_canvas_asset`` tool bound to ``store``.

    The intake wiring for binary assets: it moves image bytes the agent
    already holds (handed over by the host, produced by a separate pipeline)
    onto the canvas under ``assets/``, where every canvas file can reference
    them by relative path. It creates nothing — image *generation* belongs to
    the adopter's own tools, not the canvas core. Kept separate from
    :func:`create_canvas_tools` so the four standard tools stay a stable
    contract.
    """

    @tool
    def write_canvas_asset(
        path: str, content_base64: str, description: str, runtime: ToolRuntime
    ) -> str:
        """Store an image the canvas can reference, under `assets/`.

        `path` is a file name like `logo.png` (stored as `assets/logo.png`)
        or an explicit `assets/...` path. `content_base64` is the raw image
        bytes, base64-encoded — this tool stores bytes you already have; it
        does not create or fetch images. `description` is the version-history
        entry.

        Reference the stored file from canvas files by its relative path:
        `<img src="assets/logo.png">` in an .html page,
        `![logo](assets/logo.png)` in a document, `src: "assets/logo.png"`
        on a slide image element. The canvas displays it live and exports
        inline the bytes into the exported file. The user's uploads under
        `sources/` are referenced the same way — never copy them here.
        """
        canvas_id = _canvas_id(runtime)
        path = path.strip()
        if path.startswith(ASSETS_PREFIX):
            pass
        elif "/" not in path and path:
            path = ASSETS_PREFIX + path
        elif path.startswith(_SOURCES_PREFIX):
            return (
                "Error: files under sources/ are the user's uploads. Reference "
                'them directly (<img src="sources/...">) instead of copying.'
            )
        else:
            return (
                f"Error: assets live under {ASSETS_PREFIX} — pass a file name "
                f"or an {ASSETS_PREFIX}... path (got {path!r})."
            )
        dot = path.rfind(".")
        suffix = path[dot:].lower() if dot != -1 else ""
        if suffix not in ASSET_IMAGE_MIME:
            supported = ", ".join(sorted(ASSET_IMAGE_MIME))
            return (
                f"Error: {suffix or 'no extension'} is not an embeddable image "
                f"type ({supported}). For text content use write_canvas."
            )
        # Models often hand back a full data: URI — accept it, keep the bytes.
        encoded = content_base64.strip()
        if encoded.startswith("data:"):
            encoded = encoded.partition(",")[2]
        try:
            data = base64.b64decode(encoded, validate=True)
        except (binascii.Error, ValueError):
            return "Error: content_base64 is not valid base64."
        if not data:
            return "Error: content_base64 decoded to zero bytes."
        try:
            commit = store.write_bytes(canvas_id, path, data, description, actor="agent")
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        return (
            f"Wrote {path} ({len(data)} bytes, revision {commit.revision}). "
            f'Reference it by relative path, e.g. <img src="{path}">.'
        )

    return write_canvas_asset


_EXPECT_PATTERN = re.compile(r"^\s*(?P<key>[^\[\]=]+)\[(?P<row>\d+)\]\s*=\s*(?P<value>.*?)\s*$")
_FUNCTION_NAME_PATTERN = re.compile(r"([A-Z][A-Z0-9.]*)\s*\(")


def _values_equal(got: object, want: str) -> bool:
    """Loose comparison for expect assertions: numeric when both parse."""
    try:
        return float(str(got)) == float(want)
    except ValueError:
        return str(got).strip() == want


def _error_hint(formula: str) -> str:
    """Why a formula likely failed, steering toward supported classics."""
    unsupported = sorted(
        name
        for name in set(_FUNCTION_NAME_PATTERN.findall(formula.upper()))
        if name not in SUPPORTED_FORMULA_FUNCTIONS
    )
    if unsupported:
        return (
            f"{', '.join(unsupported)} is not supported on this canvas — rewrite with "
            "classic equivalents (SUMIFS, MATCH, TEXTJOIN, ...; see the supported list "
            "in this tool's description)"
        )
    return "the formula failed to evaluate — check references and argument types"


def create_check_table_tool(
    store: CanvasStore, *, evaluator: Sequence[str] | None = None
) -> Any:
    """Build a ``check_table`` tool bound to ``store``.

    Closes the write → check → fix loop for table formulas: the tool
    evaluates every formula cell of a ``.table.json`` file and reports the
    results — and errors — back as text the agent can act on.

    ``evaluator`` is the command that does the evaluating: an argv sequence
    that reads a ``{"columns": ..., "rows": ...}`` JSON payload on stdin and
    writes ``{"results": [...]}`` on stdout. The reference command is
    ``("node", "<canvas-react>/dist/formula-cli.js")`` — a one-shot
    subprocess around the *same* engine and function registrations the
    client uses to display formula results, so the check can never drift
    from what the canvas shows. With ``evaluator=None`` (or a missing
    runtime) the tool stays mounted but answers with honest guidance
    instead of a false verdict.
    """

    @tool
    def check_table(
        path: str, runtime: ToolRuntime, expect: list[str] | None = None
    ) -> str:
        """Evaluate the formulas of a .table.json file and report every result.

        Run this after writing or editing a table that contains formulas
        ("=..." cell values). It computes each formula with the same engine
        the canvas uses, so what it reports is what the user sees. Fix every
        ERROR (read_canvas + edit_canvas) and re-check until it reports
        0 errors. ``expect`` optionally asserts computed values, one entry
        per assertion in the form ``column_key[row_index]=value`` with the
        0-based index into ``rows`` — for example ``["total[2]=1250"]``.
        """
        canvas_id = _canvas_id(runtime)
        if not path.endswith(".table.json"):
            return f"Error: check_table reads .table.json files (got {path})."
        try:
            content = store.read(canvas_id, path).content
        except CanvasFileNotFoundError as exc:
            return f"Error: {exc}. Use list_canvas_files to see available files."

        try:
            envelope = json.loads(content)
            data = envelope.get("data") or {}
            columns = data.get("columns") or []
            rows = data.get("rows") or []
        except (ValueError, AttributeError):
            return f"Error: {path} does not contain valid table JSON."

        typed = sum(
            1
            for sheet in data.get("sheet") or []
            for cell in sheet.get("celldata") or []
            if isinstance(cell.get("v"), dict) and cell["v"].get("f")
        )

        formula_cells = sum(
            1
            for row in rows
            for value in row.values()
            if isinstance(value, str) and value.startswith("=")
        )
        # Grid formulas (the sheet editor state — where write_table_cells and
        # typing put them) are checked too: "they evaluate in the grid" was
        # assumed here once, and an agent-written formula did not.
        grid_lines: list[str] = []
        grid_errors = 0
        grid_checked = 0
        if typed and evaluator is not None:
            grid_payload = {
                "sheets": [
                    {"celldata": sheet.get("celldata") or []}
                    for sheet in data.get("sheet") or []
                ]
            }
            grid_results = _run_formula_evaluator(evaluator, grid_payload)
            if not grid_results:
                grid_lines.append(
                    f"note: {typed} grid formula(s) NOT verified — the evaluator "
                    "could not run or predates grid mode"
                )
            else:
                from .table_outline import _a1

                grid_checked = len(grid_results)
                for cell in grid_results:
                    address = f"s{cell['sheet']}!{_a1(int(cell['r']), int(cell['c']))}"
                    value = cell.get("value")
                    if value in (None, "#ERR"):
                        grid_errors += 1
                        grid_lines.append(
                            f"ERROR {address}: {cell['formula']} -> #ERR — "
                            + _error_hint(str(cell["formula"]))
                        )
                    else:
                        grid_lines.append(f"ok    {address}: {cell['formula']} -> {value}")
        elif typed:
            grid_lines.append(
                f"note: {typed} grid formula(s) NOT verified — no evaluator configured"
            )
        if formula_cells == 0 and not expect:
            if grid_checked or grid_lines:
                return "\n".join(
                    [
                        f"{grid_errors} ERROR — {grid_checked} grid formula cell(s) "
                        f"evaluated in {path}, none in rows.",
                        *grid_lines,
                    ]
                )
            return f"0 ERROR — no formula cells in {path} rows."

        if evaluator is None:
            return (
                "Error: check_table has no formula evaluator configured. The host "
                "must pass `evaluator` to create_check_table_tool (normally "
                "('node', '<canvas-react>/dist/formula-cli.js')). The formulas were "
                "NOT verified."
            )
        payload = json.dumps({"columns": columns, "rows": rows})
        try:
            proc = subprocess.run(  # noqa: S603 — host-configured argv, no shell
                list(evaluator), input=payload.encode(), capture_output=True, timeout=60
            )
            output = json.loads(proc.stdout.decode())
            results = output["results"]
        except (OSError, ValueError, KeyError, subprocess.TimeoutExpired) as exc:
            return (
                f"Error: the formula evaluator could not run ({exc}). The formulas "
                "were NOT verified — is Node.js installed and the canvas package built?"
            )

        lines: list[str] = []
        errors = 0
        computed: dict[tuple[str, int], object] = {}
        for cell in results:
            key, row_idx = str(cell["key"]), int(cell["row"])
            value = cell.get("value")
            computed[(key, row_idx)] = value
            if value == "#ERR":
                errors += 1
                lines.append(
                    f"ERROR {key}[{row_idx}]: {cell['formula']} -> #ERR — "
                    + _error_hint(str(cell["formula"]))
                )
            else:
                lines.append(f"ok    {key}[{row_idx}]: {cell['formula']} -> {value}")

        for assertion in expect or []:
            match = _EXPECT_PATTERN.match(assertion)
            if not match:
                errors += 1
                lines.append(
                    f"ERROR expect {assertion!r}: not in the form column_key[row_index]=value"
                )
                continue
            key, row_idx = match["key"].strip(), int(match["row"])
            got = computed.get((key, row_idx))
            if got is None:
                got = rows[row_idx].get(key) if row_idx < len(rows) else None
            if _values_equal(got, match["value"]):
                lines.append(f"ok    expect {key}[{row_idx}] = {match['value']}")
            else:
                errors += 1
                lines.append(f"ERROR expect {key}[{row_idx}] = {match['value']}, got {got}")

        errors += grid_errors
        checked = f"{len(results)} formula cell(s)"
        if grid_checked:
            checked += f" + {grid_checked} grid formula cell(s)"
        summary = f"{errors} ERROR — {checked} evaluated in {path}."
        return "\n".join([summary, *lines, *grid_lines])

    check_table.description += "\n\n" + formula_guidance()
    return check_table


def create_deck_tools(
    store: CanvasStore, *, converters: list[SourceConverter] | None = None
) -> list[Any]:
    """Build the tool that makes an uploaded deck editable.

    An uploaded ``.pptx`` already shows as slides — the upload path reads it
    with :func:`~langchain_canvas.pptx_import.pptx_to_slides`. But it shows
    from under ``sources/``, where the user's originals are read-only, and
    under its own name, which no exporter matches. So it can be looked at and
    nothing else.

    This copies it out: the same slides, written to a ``.slides.json`` at the
    canvas root, which edits and exports like any deck the agent builds. The
    copy names the original as its ``template``, so exporting rebuilds on the
    real masters and layouts rather than a blank page.

    Kept out of :func:`create_canvas_tools` so the four standard tools stay a
    stable contract; mount this when your agent should edit decks people send.
    ``converters`` is the list ``create_canvas_tools`` takes; a page renderer
    among them is what lets charts come across as pictures.
    """
    active_converters = default_converters() if converters is None else converters

    @tool
    def open_deck_for_editing(
        source: str,
        runtime: ToolRuntime,
        destination: str | None = None,
    ) -> str:
        """Copy an uploaded PowerPoint file into an editable canvas deck.

        `source` is the uploaded `.pptx` (usually under `sources/`). The copy
        lands at the canvas root as a `.slides.json`; pass `destination` to
        name it yourself. The upload stays where it is, unchanged.

        The copy keeps each slide's shapes, text, fonts, colours, positions
        and speaker notes, and points at the original as its template so an
        export rebuilds on the original's masters. Its pictures are stored
        under `assets/<deck name>/` and referenced by path, so the copy
        stays small enough to read whole. Tables, charts and grouped shapes
        do not come across — read the upload itself to see those.

        Text boxes that grow with their text in the original keep doing so
        (`autofit: "shape"`, shown as `grows` in the outline); ones that
        shrink their type keep that too (`"text"`, shown as `shrinks`). The
        rest are fixed: words that run longer than the box show up as an
        overflow in the deck check. Where new words do not fit a fixed box,
        shorten them, set `autofit`, or ask the user which they prefer.

        This is how an uploaded deck is revised: read the copy, then change
        its words with `set_slide_texts`, one slide per call. Changing the
        words keeps the look; geometry and styling go through `edit_canvas`.
        Do not write a new deck from scratch for a revision — it would carry
        none of the original's styling.
        """
        canvas_id = _canvas_id(runtime)
        if not source.lower().endswith(".pptx"):
            return (
                f"Error: this opens .pptx files (got {source}). "
                + _other_ways_to_open(source, default_converters())
            )
        target = (destination or _deck_copy_name(source)).strip()
        if target.startswith(_SOURCES_PREFIX):
            return (
                "Error: sources/ holds the user's uploads — put the copy "
                "somewhere else (the default is the deck's name at the canvas root)."
            )
        if not target.endswith(SLIDES_SUFFIX):
            return f"Error: the copy has to be a {SLIDES_SUFFIX} file (got {target})."
        if any(info.path == target for info in store.list_files(canvas_id)):
            return (
                f"Error: {target} is already on the canvas. Edit that one, or pass "
                "`destination` to copy under another name."
            )
        try:
            got = store.read_bytes(canvas_id, source)
        except CanvasFileNotFoundError as exc:
            return f"Error: {exc}. Use list_canvas_files to see available files."
        except CanvasStoreError as exc:
            return f"Error: {exc}."

        from .pptx_import import (
            PptxImportError,
            master_has_content,
            pptx_to_slides,
        )

        try:
            deck = pptx_to_slides(got.data)
        except PptxImportError as exc:
            return f"Error: {exc}."
        deck["template"] = source
        try:
            pictures = _extract_pictures(canvas_id, source, target, deck)
            charts, charts_dropped = _charts_as_pictures(canvas_id, source, target, deck, got.data)
            backdrops = _master_backdrops(canvas_id, target, deck, got.data)
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        try:
            commit = store.write(
                canvas_id,
                target,
                # Through the envelope encoder, never json.dumps: a .slides.json
                # without {"type","title","data"} parses as no artifact at all
                # and the canvas falls back to showing the JSON as a document.
                encode_slides(display_title(target), deck),
                f"Copy {source} for editing",
                actor="agent",
            )
        except CanvasStoreError as exc:
            return f"Error: {exc}."
        # An artifact broadcast, not a source one: source_preview_events treats
        # a .json path as a text preview, which drew the deck as its own JSON.
        writer = getattr(runtime, "stream_writer", None)
        if writer is not None:
            for event in events_for_commit(
                target,
                store.read(canvas_id, target, revision=commit.revision).content,
                is_new=True,
                revision=commit.revision,
                description=commit.description,
            ):
                writer(event)
        count = len(deck.get("slides", []))
        tables = sum(
            1
            for slide in deck.get("slides", [])
            for element in slide.get("elements", [])
            if element.get("type") == "table"
        )
        grows = sum(
            1
            for slide in deck.get("slides", [])
            for element in slide.get("elements", [])
            if element.get("type") == "text" and element.get("autofit") == "shape"
        )
        shrinks = sum(
            1
            for slide in deck.get("slides", [])
            for element in slide.get("elements", [])
            if element.get("type") == "text" and element.get("autofit") == "text"
        )
        extras = []
        if tables:
            extras.append(f"{tables} table(s) as table elements (rows editable)")
        if grows:
            extras.append(f"{grows} text box(es) grow with their text (autofit: shape)")
        if shrinks:
            extras.append(f"{shrinks} text box(es) shrink their type to fit (autofit: text)")
        if backdrops:
            extras.append(
                f"the master's own logo/footer drawn behind {backdrops} slide(s) "
                "(display only — not editable, exports from the original)"
            )
        elif master_has_content(got.data):
            extras.append(
                "the master carries a logo/footer this copy cannot draw here; "
                "it is safe and returns on export"
            )
        if charts:
            extras.append(f"{charts} chart(s) as pictures (not editable)")
        if charts_dropped:
            extras.append(f"{charts_dropped} chart(s) dropped — no page renderer to draw them")
        carried = f" Also: {'; '.join(extras)}." if extras else ""
        return (
            f"Copied {source} to {target} ({count} slide(s), {pictures} picture(s) "
            f"under {ASSETS_PREFIX}{_deck_stem(target)}/, revision {commit.revision})."
            f"{carried} Read {target}, then change its words with set_slide_texts, one slide "
            "per call — the fonts, colours and positions came from the original, so "
            "changing the words keeps the look. A box that grows takes the height "
            "its words need (mind "
            "the page bottom and what sits below it); a fixed box does not — when "
            "new words run longer than its placeholder, shorten them, set `autofit`, "
            f"or ask the user which they prefer. Export it to pptx when done; {source} "
            "keeps the user's original."
        )

    def _master_backdrops(
        canvas_id: str, target: str, deck: dict[str, Any], original: bytes
    ) -> int:
        """Give each slide the master/layout as a display-only backdrop.

        The logo and footer live on the master, out of reach on the slide —
        the editable copy used to lose them. With a page renderer mounted the
        deck is rendered once with every slide's own shapes removed, so each
        page shows exactly what the master and layout draw; identical pages
        (same layout) share one ``assets/`` file. The pptx exporter ignores
        the field — the template skin carries the real master. Returns how
        many slides got a backdrop; without a renderer, master content, or a
        clean render, none do and the copy behaves as before.
        """
        from .pptx_import import blank_slides_pptx, master_has_content

        slides = deck.get("slides", [])
        if not slides or not master_has_content(original):
            return 0
        renderer = _renderer_for(f"{_deck_stem(target)}.pptx", active_converters)
        if renderer is None:
            return 0
        blank = blank_slides_pptx(original)
        if blank is None:
            return 0
        try:
            converted = renderer.render_pages(
                blank, path=f"{_deck_stem(target)}.pptx", pages=list(range(1, len(slides) + 1))
            )
        except Exception:  # noqa: BLE001 — the backdrop is a bonus, never a failure
            return 0
        images = [block for block in converted.blocks if block.get("type") == "image"]
        if len(images) < len(slides):
            return 0
        folder = f"{ASSETS_PREFIX}{_deck_stem(target)}/"
        stored: dict[str, str] = {}
        given = 0
        for slide, block in zip(slides, images, strict=False):
            try:
                data = base64.b64decode(block["data"])
            except (KeyError, ValueError, binascii.Error):
                continue
            digest = hashlib.sha256(data).hexdigest()[:12]
            path = stored.get(digest)
            if path is None:
                path = f"{folder}master-{digest}.png"
                store.write_bytes(
                    canvas_id, path, data, f"Master backdrop for {target}", actor="agent"
                )
                stored[digest] = path
            slide["masterImage"] = path
            given += 1
        return given

    def _charts_as_pictures(
        canvas_id: str, source: str, target: str, deck: dict[str, Any], original: bytes
    ) -> tuple[int, int]:
        """Turn each chart's box into a picture of it, from the rendered page.

        The deck model holds no chart, and a chart's data is not the point of
        a slide someone is revising — its picture is. With a page renderer
        mounted, the slide is drawn as PowerPoint would print it and the
        chart's box is cut out of that page into ``assets/``; the deck gets
        an image element in the same place. Without a renderer (or when the
        render fails) the chart is dropped as before and the reply says so.
        Returns ``(pictured, dropped)`` and strips the ``charts`` side list
        the reader left on each slide, which is not a deck field.
        """
        boxes = [
            (number, box)
            for number, slide in enumerate(deck.get("slides", []), start=1)
            for box in (slide.pop("charts", None) or [])
        ]
        if not boxes:
            return 0, 0
        renderer = _renderer_for(source, active_converters)
        if renderer is None:
            return 0, len(boxes)
        try:
            from PIL import Image
        except ImportError:
            return 0, len(boxes)
        folder = f"{ASSETS_PREFIX}{_deck_stem(target)}/"
        pictured = 0
        pages: dict[int, Any] = {}
        for number, box in boxes:
            try:
                if number not in pages:
                    converted = renderer.render_pages(original, path=source, pages=[number])
                    block = next(b for b in converted.blocks if b.get("type") == "image")
                    pages[number] = Image.open(io.BytesIO(base64.b64decode(block["data"])))
                page = pages[number]
                crop = page.crop((
                    int(page.width * box["x"] / 100),
                    int(page.height * box["y"] / 100),
                    int(page.width * (box["x"] + box["w"]) / 100),
                    int(page.height * (box["y"] + box["h"]) / 100),
                ))
                buffer = io.BytesIO()
                crop.save(buffer, format="PNG")
            except Exception:  # noqa: BLE001 - one chart's failure drops that chart only
                continue
            blob = buffer.getvalue()
            path = f"{folder}chart-s{number}-{hashlib.sha256(blob).hexdigest()[:8]}.png"
            store.write_bytes(canvas_id, path, blob, f"Chart picture from {source}", actor="agent")
            deck["slides"][number - 1].setdefault("elements", []).append(
                {"id": f"chart{pictured}", "type": "image", **box, "src": path}
            )
            pictured += 1
        return pictured, len(boxes) - pictured

    def _extract_pictures(
        canvas_id: str, source: str, target: str, deck: dict[str, Any]
    ) -> int:
        """Move the copy's inline pictures under ``assets/`` and reference them.

        The reader hands pictures over as data URIs, so a deck is complete
        on its own. Stored that way, six slides with a few photos are over a
        megabyte of base64, and a model asked to edit them reads the opening
        of an encoded JPEG and nothing after — the words it was meant to
        change never reach it. The bytes go to the store, where every canvas
        file already references pictures by path, and the deck goes back to
        being the text it is. A picture repeated across slides (a logo) is
        stored once. Returns how many pictures were stored.
        """
        folder = f"{ASSETS_PREFIX}{_deck_stem(target)}/"
        stored: dict[str, str] = {}
        for slide in deck.get("slides", []):
            for element in slide.get("elements", []):
                decoded = _decode_image_data_uri(element.get("src"))
                if decoded is None:
                    continue
                suffix, blob = decoded
                digest = hashlib.sha256(blob).hexdigest()[:12]
                path = stored.get(digest)
                if path is None:
                    path = f"{folder}{digest}{suffix}"
                    store.write_bytes(
                        canvas_id, path, blob, f"Picture from {source}", actor="agent"
                    )
                    stored[digest] = path
                element["src"] = path
        return len(stored)

    return [open_deck_for_editing]
